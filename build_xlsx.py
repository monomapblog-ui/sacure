"""
民間廃棄物収集受託案件 採算管理ブック
"""
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, numbers
)
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, CellIsRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
import datetime

# ── Palette ───────────────────────────────────────────────────
C_NAVY    = "0D2B5E"
C_BLUE    = "1A5FC8"
C_GREEN   = "2EB872"
C_LGRAY   = "F2F6FC"
C_DGRAY   = "55657A"
C_WHITE   = "FFFFFF"
C_YELLOW  = "FFF3CD"
C_RED_LT  = "FFE0E0"
C_GREEN_LT= "D4EDDA"
C_BLUE_LT = "D0E8FF"
C_BORDER  = "C5D5E8"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, size=11, color=C_NAVY, name="Yu Gothic UI"):
    return Font(bold=bold, size=size, color=color, name=name)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def thin_border(top=True, bottom=True, left=True, right=True):
    s = Side(style="thin", color=C_BORDER)
    n = Side(style=None)
    return Border(
        top=s if top else n,
        bottom=s if bottom else n,
        left=s if left else n,
        right=s if right else n,
    )

def set_cell(ws, row, col, value=None, bold=False, size=11, fg=None,
             font_color=C_NAVY, h_align="left", v_align="center",
             wrap=False, fmt=None, border=True, merge=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, size=size, color=font_color)
    c.alignment = align(h=h_align, v=v_align, wrap=wrap)
    if fg:
        c.fill = fill(fg)
    if border:
        c.border = thin_border()
    if fmt:
        c.number_format = fmt
    if merge:
        ws.merge_cells(
            start_row=row, start_column=col,
            end_row=merge[0], end_column=merge[1]
        )
    return c

# ════════════════════════════════════════════════════════════════
#  SHEET 1: 案件マスタ
# ════════════════════════════════════════════════════════════════

MASTER_SAMPLES = [
    ("A001","吉野家フーズ㈱","飲食","大田区","山田 太郎","2024-04-01","","週2回","稼働中","一般廃棄物（食品残渣）",""),
    ("A002","東急㈱ 渋谷エリア","物流","渋谷区","山田 太郎","2024-06-01","","月4回","稼働中","産業廃棄物（紙くず・木くず）",""),
    ("A003","ファミリーマート㈱","小売","練馬区","佐藤 花子","2023-10-01","","週3回","稼働中","一般廃棄物（食品廃棄物）",""),
    ("A004","永谷園HD㈱","製造","品川区","佐藤 花子","2025-01-01","","週1回","稼働中","産業廃棄物（食料残渣）",""),
    ("A005","明治大学","教育","千代田区","鈴木 次郎","2024-09-01","2026-03-31","月2回","稼働中","一般廃棄物（古紙・段ボール）","年度契約"),
    ("A006","川崎市 環境局","行政","川崎市","鈴木 次郎","2025-04-01","","週5回","稼働中","一般廃棄物（可燃）","行政委託"),
    ("A007","白洋舍㈱","サービス","大田区","田中 美紀","2024-11-01","","週2回","稼働中","産業廃棄物（繊維くず）",""),
    ("A008","レバテック㈱","IT","渋谷区","田中 美紀","2025-03-01","","月1回","稼働中","産業廃棄物（廃プラ・金属）",""),
    ("A009","カヤック㈱","IT","神奈川","田中 美紀","2024-07-01","","月2回","稼働中","一般廃棄物（オフィス系）",""),
    ("A010","河合塾","教育","大田区","山田 太郎","2024-08-01","","月3回","稼働中","一般廃棄物（紙くず）",""),
]

def build_master(ws):
    ws.title = "案件マスタ"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 36
    ws.row_dimensions[2].height = 22

    # Title row
    ws.merge_cells("A1:K1")
    c = ws["A1"]
    c.value     = "民間廃棄物収集受託案件　案件マスタ"
    c.font      = Font(bold=True, size=16, color=C_WHITE, name="Yu Gothic UI")
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="center")

    headers = [
        ("案件ID",8),("取引先名",22),("業種",10),("エリア",12),
        ("担当者",12),("契約開始日",13),("契約終了日",13),
        ("回収頻度",12),("ステータス",12),("廃棄物種別",22),("備考",18),
    ]
    for col, (hdr, width) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
        set_cell(ws, 2, col, hdr, bold=True, size=10,
                 fg=C_BLUE, font_color=C_WHITE, h_align="center")

    status_dv = DataValidation(
        type="list", formula1='"稼働中,交渉中,一時停止,終了"',
        allow_blank=True, showDropDown=False
    )
    ws.add_data_validation(status_dv)

    for r, row in enumerate(MASTER_SAMPLES, 3):
        ws.row_dimensions[r].height = 20
        fgs = [C_LGRAY if r % 2 == 0 else C_WHITE] * 11
        vals = list(row)
        for col, (val, bg) in enumerate(zip(vals, fgs), 1):
            fmt = "YYYY-MM-DD" if col in (6, 7) and val else None
            v = datetime.date.fromisoformat(val) if fmt and val else val
            set_cell(ws, r, col, v, size=10, fg=bg, h_align="center" if col in (1,3,4,5,6,7,8,9) else "left", fmt=fmt)
        status_dv.add(ws.cell(row=r, column=9))

    # Total label
    last = 2 + len(MASTER_SAMPLES) + 1
    ws.row_dimensions[last].height = 20
    set_cell(ws, last, 1, "※ 月次採算入力シートにも案件IDを使用してください",
             size=9, fg=C_LGRAY, font_color=C_DGRAY, border=False)

# ════════════════════════════════════════════════════════════════
#  SHEET 2: 月次採算入力
# ════════════════════════════════════════════════════════════════

MONTHS = [f"2026-{m:02d}" for m in range(4, 10)]  # Apr-Sep 2026
SAMPLE_DATA = {
    ("A001","2026-04"): dict(basic=185000, spot=12000, other_s=0, driver=42000, fuel=18000, vehicle=8000, process=35000, haul=0, other_c=5000),
    ("A001","2026-05"): dict(basic=185000, spot=8000,  other_s=0, driver=42000, fuel=16000, vehicle=8000, process=33000, haul=0, other_c=4000),
    ("A002","2026-04"): dict(basic=320000, spot=45000, other_s=0, driver=78000, fuel=32000, vehicle=14000, process=68000, haul=12000, other_c=8000),
    ("A003","2026-04"): dict(basic=210000, spot=0,     other_s=0, driver=55000, fuel=22000, vehicle=10000, process=48000, haul=0,    other_c=6000),
    ("A004","2026-04"): dict(basic=95000,  spot=5000,  other_s=0, driver=28000, fuel=11000, vehicle=6000,  process=22000, haul=0,    other_c=3000),
    ("A006","2026-04"): dict(basic=580000, spot=0,     other_s=0, driver=145000,fuel=62000, vehicle=28000, process=95000, haul=0,    other_c=15000),
    ("A007","2026-04"): dict(basic=148000, spot=18000, other_s=0, driver=38000, fuel=15000, vehicle=7000,  process=31000, haul=5000, other_c=4000),
}

def build_input(ws):
    ws.title = "月次採算入力"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D4"

    # Row 1: title
    ws.merge_cells("A1:S1")
    c = ws["A1"]
    c.value = "月次採算入力シート　（案件別・月別）"
    c.font  = Font(bold=True, size=15, color=C_WHITE, name="Yu Gothic UI")
    c.fill  = fill(C_NAVY)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 34

    # Row 2: section headers (merged)
    sections = [
        (1, 3, "案件情報"),
        (4, 4, "集計月"),
        (5, 8, "売　上"),
        (9, 15, "原　価"),
        (16, 19, "採　算"),
    ]
    ws.row_dimensions[2].height = 22
    ws.row_dimensions[3].height = 22
    section_colors = {
        "案件情報": C_NAVY,
        "集計月":   C_DGRAY,
        "売　上":   "1565A8",
        "原　価":   "7B3F00",
        "採　算":   "1A6B2E",
    }
    for sc, ec, label in sections:
        ws.merge_cells(start_row=2, start_column=sc, end_row=2, end_column=ec)
        c = ws.cell(row=2, column=sc, value=label)
        c.font      = Font(bold=True, size=10, color=C_WHITE, name="Yu Gothic UI")
        c.fill      = fill(section_colors[label])
        c.alignment = align(h="center")
        c.border    = thin_border()

    # Row 3: column headers
    col_headers = [
        ("案件ID",9),("取引先名",20),("担当者",12),
        ("集計年月",11),
        ("基本料金",12),("スポット料金",12),("その他売上",11),("売上合計",12),
        ("人件費",11),("燃料費",10),("車両維持費",11),("中間処理費",11),
        ("搬入・運搬費",12),("その他原価",11),("原価合計",12),
        ("粗利益",12),("粗利率",10),("採算判定",10),("備考",16),
    ]
    header_colors = (
        [C_NAVY]*3 + [C_DGRAY] +
        ["1565A8"]*4 +
        ["7B3F00"]*7 +
        ["1A6B2E"]*4
    )
    for col, ((hdr, width), fg_c) in enumerate(zip(col_headers, header_colors), 1):
        ws.column_dimensions[get_column_letter(col)].width = width
        set_cell(ws, 3, col, hdr, bold=True, size=9,
                 fg=fg_c, font_color=C_WHITE, h_align="center")

    START_ROW = 4
    JPY_FMT   = '#,##0'
    PCT_FMT   = '0.0%'

    row = START_ROW
    for proj_id, ym in [
        ("A001","2026-04"),("A001","2026-05"),
        ("A002","2026-04"),("A003","2026-04"),
        ("A004","2026-04"),("A006","2026-04"),
        ("A007","2026-04"),
    ]:
        d = SAMPLE_DATA.get((proj_id, ym), {})
        name_map = {
            "A001":"吉野家フーズ㈱","A002":"東急㈱","A003":"ファミリーマート㈱",
            "A004":"永谷園HD㈱","A006":"川崎市 環境局","A007":"白洋舎㈱",
        }
        person_map = {
            "A001":"山田 太郎","A002":"山田 太郎","A003":"佐藤 花子",
            "A004":"佐藤 花子","A006":"鈴木 次郎","A007":"田中 美紀",
        }
        bg = C_LGRAY if row % 2 == 0 else C_WHITE

        set_cell(ws, row, 1,  proj_id,              size=10, fg=bg, h_align="center")
        set_cell(ws, row, 2,  name_map.get(proj_id,""), size=10, fg=bg)
        set_cell(ws, row, 3,  person_map.get(proj_id,""), size=10, fg=bg, h_align="center")
        set_cell(ws, row, 4,  ym,                   size=10, fg=bg, h_align="center")

        basic    = d.get("basic",0)
        spot     = d.get("spot",0)
        other_s  = d.get("other_s",0)
        driver   = d.get("driver",0)
        fuel     = d.get("fuel",0)
        vehicle  = d.get("vehicle",0)
        process  = d.get("process",0)
        haul     = d.get("haul",0)
        other_c  = d.get("other_c",0)

        # 売上 E F G H
        set_cell(ws, row, 5,  basic,   size=10, fg=C_BLUE_LT, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 6,  spot,    size=10, fg=C_BLUE_LT, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 7,  other_s, size=10, fg=C_BLUE_LT, fmt=JPY_FMT, h_align="right")
        h_c = ws.cell(row=row, column=8)
        h_c.value          = f"=E{row}+F{row}+G{row}"
        h_c.number_format  = JPY_FMT
        h_c.font           = font(bold=True, size=10, color=C_NAVY)
        h_c.fill           = fill(C_BLUE_LT)
        h_c.alignment      = align(h="right")
        h_c.border         = thin_border()

        # 原価 I J K L M N
        set_cell(ws, row,  9, driver,  size=10, fg=C_YELLOW, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 10, fuel,    size=10, fg=C_YELLOW, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 11, vehicle, size=10, fg=C_YELLOW, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 12, process, size=10, fg=C_YELLOW, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 13, haul,    size=10, fg=C_YELLOW, fmt=JPY_FMT, h_align="right")
        set_cell(ws, row, 14, other_c, size=10, fg=C_YELLOW, fmt=JPY_FMT, h_align="right")
        o_c = ws.cell(row=row, column=15)
        o_c.value         = f"=I{row}+J{row}+K{row}+L{row}+M{row}+N{row}"
        o_c.number_format = JPY_FMT
        o_c.font          = font(bold=True, size=10, color=C_NAVY)
        o_c.fill          = fill(C_YELLOW)
        o_c.alignment     = align(h="right")
        o_c.border        = thin_border()

        # 採算 P Q R S
        p_c = ws.cell(row=row, column=16)
        p_c.value         = f"=H{row}-O{row}"
        p_c.number_format = JPY_FMT
        p_c.font          = font(bold=True, size=10, color=C_NAVY)
        p_c.fill          = fill(C_GREEN_LT)
        p_c.alignment     = align(h="right")
        p_c.border        = thin_border()

        q_c = ws.cell(row=row, column=17)
        q_c.value         = f"=IF(H{row}<>0,P{row}/H{row},0)"
        q_c.number_format = PCT_FMT
        q_c.font          = font(bold=True, size=10, color=C_NAVY)
        q_c.fill          = fill(C_GREEN_LT)
        q_c.alignment     = align(h="center")
        q_c.border        = thin_border()

        r_c = ws.cell(row=row, column=18)
        r_c.value         = f'=IF(Q{row}>=0.2,"◎ 良好",IF(Q{row}>=0.1,"○ 適正",IF(Q{row}>=0,"△ 要注意","× 赤字")))'
        r_c.font          = font(bold=True, size=10, color=C_NAVY)
        r_c.fill          = fill(C_GREEN_LT)
        r_c.alignment     = align(h="center")
        r_c.border        = thin_border()

        set_cell(ws, row, 19, "", size=10, fg=bg)

        row += 1

    # Empty input rows (30 more)
    for _ in range(30):
        bg = C_LGRAY if row % 2 == 0 else C_WHITE
        for col in range(1, 20):
            set_cell(ws, row, col, None, size=10, fg=bg,
                     h_align="right" if col >= 5 else ("center" if col in (1,3,4) else "left"))
        # formulas
        for col, formula in [
            (8,  f"=IF(E{row}+F{row}+G{row}=0,\"\",E{row}+F{row}+G{row})"),
            (15, f"=IF(I{row}+J{row}+K{row}+L{row}+M{row}+N{row}=0,\"\",I{row}+J{row}+K{row}+L{row}+M{row}+N{row})"),
            (16, f'=IF(OR(H{row}="",H{row}=0),"",H{row}-O{row})'),
            (17, f'=IF(OR(H{row}="",H{row}=0),"",P{row}/H{row})'),
            (18, f'=IF(Q{row}="","",IF(Q{row}>=0.2,"◎ 良好",IF(Q{row}>=0.1,"○ 適正",IF(Q{row}>=0,"△ 要注意","× 赤字"))))'),
        ]:
            c2 = ws.cell(row=row, column=col)
            c2.value         = formula
            c2.number_format = PCT_FMT if col == 17 else ('#,##0' if col in (8,15,16) else '@')
            c2.font          = font(bold=(col in (8,15,16,17,18)), size=10)
            c2.fill          = fill(C_BLUE_LT if col in (5,6,7,8) else (C_YELLOW if col in (9,10,11,12,13,14,15) else (C_GREEN_LT if col in (16,17,18) else bg)))
            c2.alignment     = align(h="center" if col in (18,) else "right")
            c2.border        = thin_border()
        row += 1

    # Conditional formatting for 判定列
    last_row = row - 1
    ws.conditional_formatting.add(
        f"R4:R{last_row}",
        FormulaRule(formula=[f'R4="◎ 良好"'],  fill=fill("C8F7D4"), font=Font(bold=True, color="145A32"))
    )
    ws.conditional_formatting.add(
        f"R4:R{last_row}",
        FormulaRule(formula=[f'R4="○ 適正"'],  fill=fill(C_GREEN_LT), font=Font(bold=True, color="1A6B2E"))
    )
    ws.conditional_formatting.add(
        f"R4:R{last_row}",
        FormulaRule(formula=[f'R4="△ 要注意"'], fill=fill(C_YELLOW),   font=Font(bold=True, color="856404"))
    )
    ws.conditional_formatting.add(
        f"R4:R{last_row}",
        FormulaRule(formula=[f'R4="× 赤字"'],  fill=fill(C_RED_LT),   font=Font(bold=True, color="8B0000"))
    )

    # Color scale for 粗利率
    ws.conditional_formatting.add(
        f"Q4:Q{last_row}",
        ColorScaleRule(
            start_type="num", start_value=0,   start_color="FF6B6B",
            mid_type="num",   mid_value=0.1,   mid_color="FFD93D",
            end_type="num",   end_value=0.3,   end_color="6BCB77",
        )
    )

# ════════════════════════════════════════════════════════════════
#  SHEET 3: 月次集計
# ════════════════════════════════════════════════════════════════

def build_monthly(ws):
    ws.title = "月次集計"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value     = "月次採算集計　（月別サマリー）"
    c.font      = Font(bold=True, size=15, color=C_WHITE, name="Yu Gothic UI")
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 22

    heads = [
        ("集計年月",14),("稼働案件数",14),("売上合計",16),("原価合計",16),
        ("粗利益合計",16),("粗利率",12),("前月比 売上",14),("前月比 粗利",14),("備考",18),
    ]
    for col, (h, w) in enumerate(heads, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        set_cell(ws, 2, col, h, bold=True, size=10, fg=C_BLUE, font_color=C_WHITE, h_align="center")

    months = [f"2026-{m:02d}" for m in range(4, 10)]
    # sample aggregated data (manual for now — in real use link to 月次採算入力)
    monthly_vals = [
        ("2026-04", 7, 1720000, 1055000),
        ("2026-05", 7, 1698000, 1038000),
        ("2026-06", 8, 1850000, 1110000),
        ("2026-07", 9, 1920000, 1132000),
        ("2026-08", 9, 1885000, 1120000),
        ("2026-09", 10,2100000, 1218000),
    ]

    for r, (ym, cnt, sales, cost) in enumerate(monthly_vals, 3):
        ws.row_dimensions[r].height = 22
        bg = C_LGRAY if r % 2 == 0 else C_WHITE
        gross  = sales - cost
        rate   = gross / sales if sales else 0

        set_cell(ws, r, 1, ym,    size=11, fg=bg, h_align="center")
        set_cell(ws, r, 2, cnt,   size=11, fg=bg, h_align="center", fmt="#,##0")
        set_cell(ws, r, 3, sales, size=11, fg=bg, h_align="right",  fmt="#,##0")
        set_cell(ws, r, 4, cost,  size=11, fg=bg, h_align="right",  fmt="#,##0")
        set_cell(ws, r, 5, gross, size=11, fg=bg, h_align="right",  fmt="#,##0", bold=True)
        set_cell(ws, r, 6, rate,  size=11, fg=bg, h_align="center", fmt="0.0%",  bold=True)

        # MoM formulas
        if r == 3:
            set_cell(ws, r, 7, "—", size=10, fg=bg, h_align="center")
            set_cell(ws, r, 8, "—", size=10, fg=bg, h_align="center")
        else:
            prev = r - 1
            c7 = ws.cell(row=r, column=7)
            c7.value         = f"=IF(C{prev}<>0,C{r}/C{prev}-1,0)"
            c7.number_format = "+0.0%;-0.0%;0.0%"
            c7.font          = font(size=10)
            c7.fill          = fill(bg)
            c7.alignment     = align(h="center")
            c7.border        = thin_border()

            c8 = ws.cell(row=r, column=8)
            c8.value         = f"=IF(E{prev}<>0,E{r}/E{prev}-1,0)"
            c8.number_format = "+0.0%;-0.0%;0.0%"
            c8.font          = font(size=10)
            c8.fill          = fill(bg)
            c8.alignment     = align(h="center")
            c8.border        = thin_border()

        set_cell(ws, r, 9, "", size=10, fg=bg)

    # Total row
    last_data = 3 + len(monthly_vals) - 1
    tr = last_data + 1
    ws.row_dimensions[tr].height = 24
    set_cell(ws, tr, 1, "合　計", bold=True, size=11, fg=C_NAVY, font_color=C_WHITE, h_align="center")
    for col, col_letter in [(2,"B"),(3,"C"),(4,"D"),(5,"E")]:
        tc = ws.cell(row=tr, column=col)
        tc.value         = f"=SUM({col_letter}3:{col_letter}{last_data})"
        tc.number_format = "#,##0" if col != 2 else "#,##0"
        tc.font          = Font(bold=True, size=11, color=C_WHITE, name="Yu Gothic UI")
        tc.fill          = fill(C_NAVY)
        tc.alignment     = align(h="right")
        tc.border        = thin_border()
    tc6 = ws.cell(row=tr, column=6)
    tc6.value         = f"=IF(C{tr}<>0,E{tr}/C{tr},0)"
    tc6.number_format = "0.0%"
    tc6.font          = Font(bold=True, size=11, color=C_WHITE, name="Yu Gothic UI")
    tc6.fill          = fill(C_NAVY)
    tc6.alignment     = align(h="center")
    tc6.border        = thin_border()
    for col in [7,8,9]:
        set_cell(ws, tr, col, "", fg=C_NAVY, bold=True, font_color=C_WHITE)

    # Chart: 売上・原価・粗利 bar chart
    chart = BarChart()
    chart.type   = "col"
    chart.title  = "月別 売上 / 原価 / 粗利益"
    chart.y_axis.title = "金額（円）"
    chart.x_axis.title = "月"
    chart.width  = 22
    chart.height = 12
    chart.grouping = "clustered"

    months_ref = Reference(ws, min_col=1, min_row=3, max_row=last_data)
    for col_idx, label in [(3,"売上"),(4,"原価"),(5,"粗利益")]:
        data = Reference(ws, min_col=col_idx, min_row=2, max_row=last_data)
        chart.add_data(data, titles_from_data=True)
    chart.set_categories(months_ref)
    chart.series[0].graphicalProperties.solidFill = C_BLUE
    chart.series[1].graphicalProperties.solidFill = "D4A800"
    chart.series[2].graphicalProperties.solidFill = C_GREEN
    ws.add_chart(chart, "A13")

# ════════════════════════════════════════════════════════════════
#  SHEET 4: 採算ダッシュボード（案件別）
# ════════════════════════════════════════════════════════════════

def build_dashboard(ws):
    ws.title = "採算ダッシュボード"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "採算ダッシュボード　案件別ランキング（直近月）"
    c.font      = Font(bold=True, size=15, color=C_WHITE, name="Yu Gothic UI")
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 34
    ws.row_dimensions[2].height = 22

    heads = [
        ("順位",8),("案件ID",10),("取引先名",22),("業種",10),("集計月",11),
        ("売上合計",14),("原価合計",14),("粗利益",14),("粗利率",11),("採算判定",12),
    ]
    for col, (h, w) in enumerate(heads, 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        set_cell(ws, 2, col, h, bold=True, size=10, fg=C_BLUE, font_color=C_WHITE, h_align="center")

    dash_data = [
        (1,"A006","川崎市 環境局","行政","2026-04",580000,345000,235000),
        (2,"A002","東急㈱","物流","2026-04",365000,212000,153000),
        (3,"A001","吉野家フーズ㈱","飲食","2026-04",197000,108000,89000),
        (4,"A007","白洋舎㈱","サービス","2026-04",166000,100000,66000),
        (5,"A003","ファミリーマート㈱","小売","2026-04",210000,141000,69000),
        (6,"A004","永谷園HD㈱","製造","2026-04",100000,70000,30000),
        (7,"A008","レバテック㈱","IT","2026-04",55000,41000,14000),
    ]

    rank_colors = ["FFD700","C0C0C0","CD7F32","","","",""]

    for r, (rank, pid, name, industry, ym, sales, cost, gross) in enumerate(dash_data, 3):
        ws.row_dimensions[r].height = 22
        bg = C_LGRAY if r % 2 == 0 else C_WHITE
        rate = gross / sales if sales else 0

        if rank <= 3:
            rank_bg = rank_colors[rank-1]
        else:
            rank_bg = bg

        set_cell(ws, r, 1,  rank,     size=12, bold=(rank<=3), fg=rank_bg, h_align="center")
        set_cell(ws, r, 2,  pid,      size=10, fg=bg, h_align="center")
        set_cell(ws, r, 3,  name,     size=10, fg=bg)
        set_cell(ws, r, 4,  industry, size=10, fg=bg, h_align="center")
        set_cell(ws, r, 5,  ym,       size=10, fg=bg, h_align="center")
        set_cell(ws, r, 6,  sales,    size=10, fg=bg, fmt="#,##0", h_align="right")
        set_cell(ws, r, 7,  cost,     size=10, fg=bg, fmt="#,##0", h_align="right")
        set_cell(ws, r, 8,  gross,    size=11, bold=True, fg=bg, fmt="#,##0", h_align="right")
        set_cell(ws, r, 9,  rate,     size=11, bold=True, fg=bg, fmt="0.0%", h_align="center")

        verdict = "◎ 良好" if rate >= 0.2 else ("○ 適正" if rate >= 0.1 else ("△ 要注意" if rate >= 0 else "× 赤字"))
        v_color = {"◎ 良好":"C8F7D4","○ 適正":C_GREEN_LT,"△ 要注意":C_YELLOW,"× 赤字":C_RED_LT}
        f_color = {"◎ 良好":"145A32","○ 適正":"1A6B2E","△ 要注意":"856404","× 赤字":"8B0000"}
        set_cell(ws, r, 10, verdict, size=11, bold=True,
                 fg=v_color[verdict], font_color=f_color[verdict], h_align="center")

    # Summary note
    note_row = 3 + len(dash_data) + 1
    ws.merge_cells(f"A{note_row}:J{note_row}")
    n = ws.cell(row=note_row, column=1)
    n.value     = "判定基準：◎ 粗利率20%以上　○ 10〜20%　△ 0〜10%　× マイナス（赤字）"
    n.font      = Font(size=9, italic=True, color=C_DGRAY, name="Yu Gothic UI")
    n.fill      = fill(C_LGRAY)
    n.alignment = align(h="left")
    n.border    = thin_border()

    # Pie chart: 売上構成
    pie_data = [
        ("川崎市 環境局", 580000),
        ("東急㈱",       365000),
        ("吉野家フーズ㈱",197000),
        ("ファミリーマート㈱",210000),
        ("その他",       210000),
    ]
    # write pie source data
    ws.row_dimensions[note_row+2].height = 18
    set_cell(ws, note_row+2, 1, "取引先", bold=True, size=10, fg=C_BLUE, font_color=C_WHITE, h_align="center")
    set_cell(ws, note_row+2, 2, "売上",   bold=True, size=10, fg=C_BLUE, font_color=C_WHITE, h_align="center")
    for i, (name, val) in enumerate(pie_data):
        r2 = note_row + 3 + i
        set_cell(ws, r2, 1, name, size=10, fg=C_LGRAY)
        set_cell(ws, r2, 2, val,  size=10, fg=C_LGRAY, fmt="#,##0", h_align="right")

    from openpyxl.chart import PieChart
    pie = PieChart()
    pie.title  = "案件別 売上構成"
    pie.width  = 16
    pie.height = 12
    labels = Reference(ws, min_col=1, min_row=note_row+3, max_row=note_row+2+len(pie_data))
    data   = Reference(ws, min_col=2, min_row=note_row+2, max_row=note_row+2+len(pie_data))
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    ws.add_chart(pie, f"D{note_row+2}")

# ════════════════════════════════════════════════════════════════
#  SHEET 5: 使い方ガイド
# ════════════════════════════════════════════════════════════════

def build_guide(ws):
    ws.title = "使い方ガイド"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 52

    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value     = "民間廃棄物収集受託案件　採算管理ブック　使い方ガイド"
    c.font      = Font(bold=True, size=16, color=C_WHITE, name="Yu Gothic UI")
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="center")
    ws.row_dimensions[1].height = 36

    steps = [
        ("STEP 1", "案件マスタに登録",
         "新しい契約案件が発生したら「案件マスタ」シートに案件IDと基本情報を登録してください。\n"
         "案件IDは「A001」形式で通し番号管理を推奨します。"),
        ("STEP 2", "月次採算を入力",
         "毎月末に「月次採算入力」シートへ各案件の実績を入力します。\n"
         "売上（基本料金・スポット等）と原価（人件費・燃料費・処理費等）を入力すると\n"
         "粗利益・粗利率・採算判定が自動で計算されます。"),
        ("STEP 3", "月次集計を確認",
         "「月次集計」シートで月ごとの売上・原価・粗利益の推移を確認します。\n"
         "グラフで月別トレンドを可視化しています。"),
        ("STEP 4", "ダッシュボードで状況把握",
         "「採算ダッシュボード」シートで案件別の採算ランキングを確認します。\n"
         "採算判定が「△ 要注意」「× 赤字」の案件は早期に改善策を検討してください。"),
    ]

    row = 3
    for step, title, desc in steps:
        ws.row_dimensions[row].height = 24
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1)
        c.value     = f"  {step}  {title}"
        c.font      = Font(bold=True, size=12, color=C_WHITE, name="Yu Gothic UI")
        c.fill      = fill(C_BLUE)
        c.alignment = align(h="left")
        row += 1

        for line in desc.split("\n"):
            ws.row_dimensions[row].height = 20
            ws.merge_cells(f"B{row}:C{row}")
            c2 = ws.cell(row=row, column=2)
            c2.value     = line
            c2.font      = Font(size=11, color=C_NAVY, name="Yu Gothic UI")
            c2.fill      = fill(C_LGRAY)
            c2.alignment = align(h="left", v="center")
            row += 1
        row += 1

    # 採算判定基準
    ws.row_dimensions[row].height = 28
    ws.merge_cells(f"A{row}:C{row}")
    c = ws.cell(row=row, column=1)
    c.value     = "  採算判定基準"
    c.font      = Font(bold=True, size=12, color=C_WHITE, name="Yu Gothic UI")
    c.fill      = fill(C_NAVY)
    c.alignment = align(h="left")
    row += 1

    criteria = [
        ("◎ 良好",   "粗利率 20%以上",    "C8F7D4", "145A32"),
        ("○ 適正",   "粗利率 10〜20%",    C_GREEN_LT, "1A6B2E"),
        ("△ 要注意", "粗利率 0〜10%",     C_YELLOW,  "856404"),
        ("× 赤字",   "粗利率 マイナス",   C_RED_LT,  "8B0000"),
    ]
    for verdict, crit, bg_c, fc in criteria:
        ws.row_dimensions[row].height = 22
        ws.column_dimensions["B"].width = 28
        c3 = ws.cell(row=row, column=2, value=verdict)
        c3.font      = Font(bold=True, size=11, color=fc, name="Yu Gothic UI")
        c3.fill      = fill(bg_c)
        c3.alignment = align(h="center")
        c3.border    = thin_border()
        c4 = ws.cell(row=row, column=3, value=crit)
        c4.font      = Font(size=11, color=C_NAVY, name="Yu Gothic UI")
        c4.fill      = fill(bg_c)
        c4.alignment = align(h="left")
        c4.border    = thin_border()
        row += 1

# ════════════════════════════════════════════════════════════════
#  BUILD
# ════════════════════════════════════════════════════════════════

def build():
    wb = Workbook()
    # Remove default sheet
    default = wb.active
    wb.remove(default)

    ws1 = wb.create_sheet()
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()

    build_master(ws1)
    build_input(ws2)
    build_monthly(ws3)
    build_dashboard(ws4)
    build_guide(ws5)

    # Tab colors
    ws1.sheet_properties.tabColor = C_NAVY
    ws2.sheet_properties.tabColor = "1565A8"
    ws3.sheet_properties.tabColor = C_GREEN
    ws4.sheet_properties.tabColor = "E07B00"
    ws5.sheet_properties.tabColor = C_DGRAY

    out = "/home/user/sacure/sacure_採算管理_民間廃棄物収集.xlsx"
    wb.save(out)
    print(f"Saved: {out}")

if __name__ == "__main__":
    build()
