"""
採算管理ブック v4 — ルート別・案件別の採算確認に特化（4シート）
"""
import openpyxl, xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, Reference

SRC_MASTER = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/83a04c7f-__________________.xlsx"
SRC_2026   = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/912c64bc-___2026.xlsx"
SRC_R76    = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/8a0a54a8-__R7.6.xls"

# ── カラー ─────────────────────────────────────────────────────
NAVY  = "0D2B5E";  BLUE  = "1A5FC8";  GREEN = "2EB872"
GRAY  = "55657A";  WHITE = "FFFFFF";  LGRAY = "F2F6FC"
YEL   = "FFF3CD";  REDLT = "FFE0E0";  GRNLT = "D4EDDA"
BLULT = "D0E8FF";  PURLT = "EDE7F6";  BDR_C = "C5D5E8"

# 案件カラー
CASE = {
    "C001": {"name":"東急グループ案件", "bg":BLULT, "hd":"1565A8"},
    "C002": {"name":"都営地下鉄案件",   "bg":PURLT, "hd":"5B2D8E"},
    "C003": {"name":"一般事業者案件",   "bg":GRNLT, "hd":"1A6B2E"},
}

# ルート定義
ROUTES = [
    ("612-1","運送（本社）",   "NAVY"),
    ("612-2","入札（公共）",   "NAVY"),
    ("613-1","車輌",           "BLUE"),
    ("613-2","KC",             "BLUE"),
    ("614-1","コンテナ",       "GREEN"),
    ("614-2","ENDO",           "GREEN"),
    ("614-3","現金",           "GREEN"),
    ("615-1","板橋",           "NAVY"),
    ("615-2","新横浜",         "NAVY"),
    ("615-3","現金2",          "BLUE"),
    ("615-4","有価物",         "GREEN"),
]
CODES = [r[0] for r in ROUTES]

# 配分設定 (C001%, C002%, C003%)
MAPPING = {
    "612-1": (0,   0, 100),
    "612-2": (0, 100,   0),
    "613-1": (20,  10,  70),
    "613-2": (0,   0,   0),
    "614-1": (0,   0, 100),
    "614-2": (0,   0, 100),
    "614-3": (0,   0, 100),
    "615-1": (60,  0,  40),
    "615-2": (0,   0, 100),
    "615-3": (0,   0, 100),
    "615-4": (0,   0, 100),
}

def fill(h): return PatternFill("solid", fgColor=h)
def f(bold=False, sz=11, color=NAVY, name="Yu Gothic UI"):
    return Font(bold=bold, size=sz, color=color, name=name)
def a(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def b():
    s = Side(style="thin", color=BDR_C)
    return Border(top=s, bottom=s, left=s, right=s)

def c(ws, row, col, val=None, bold=False, sz=11, bg=None, fc=NAVY,
      ha="left", va="center", wrap=False, fmt=None):
    cell = ws.cell(row=row, column=col, value=val)
    cell.font = f(bold=bold, sz=sz, color=fc)
    cell.alignment = a(h=ha, v=va, wrap=wrap)
    cell.border = b()
    if bg:  cell.fill = fill(bg)
    if fmt: cell.number_format = fmt
    return cell

# ────────────────────────────────────────────────────────────────
#  データ取得
# ────────────────────────────────────────────────────────────────
def reiwa(s):
    s = str(s).strip()
    if not s.startswith("R"): return None
    try:
        era, m = s[1:].split(".")
        return f"{int(era)+2018}-{int(m):02d}"
    except: return None

def load_sales():
    sales = {}

    def parse_summary(rows, header):
        for row in rows:
            if not row or not row[0] or not str(row[0]).startswith("R"): continue
            ym = reiwa(str(row[0]))
            if not ym: continue
            d = {}
            for code in CODES:
                for ci, h in enumerate(header):
                    if str(h).strip() == code and ci < len(row):
                        try: d[code] = float(row[ci]) if row[ci] else 0.0
                        except: d[code] = 0.0
                        break
            for ci, h in enumerate(header):
                if "売" in str(h) and "上" in str(h) and ci < len(row):
                    try: d["_total"] = float(row[ci]) if row[ci] else 0.0
                    except: d["_total"] = 0.0
                    break
            if any(v > 0 for v in d.values()):
                sales[ym] = d

    def parse_month_header(header):
        col_ym = {}
        cur_era = None
        for ci, h in enumerate(header):
            if not h: continue
            h = str(h).strip()
            if h.startswith("R") and "." in h:
                try:
                    era_s, m_s = h[1:].split(".")
                    cur_era = int(era_s)
                    col_ym[ci] = f"{cur_era+2018}-{int(float(m_s)):02d}"
                except: pass
            elif cur_era is not None:
                try:
                    m = int(float(h))
                    if 1 <= m <= 12:
                        col_ym[ci] = f"{cur_era+2018}-{m:02d}"
                except: pass
        return col_ym

    # R7.6.xls: R6.7〜R7.6
    wb = xlrd.open_workbook(SRC_R76)
    ws = wb.sheet_by_name("売上")
    hdr = [str(ws.cell_value(2, c)).strip() for c in range(ws.ncols)]
    rows = [[ws.cell_value(r, c) for c in range(ws.ncols)] for r in range(3, ws.nrows)]
    parse_summary(rows, hdr)

    # 2026.xlsx 売上シート: R7.7, R7.8（全11ルート確定値）
    wb2 = openpyxl.load_workbook(SRC_2026, data_only=True)
    ws2 = wb2["売上"]
    hdr2 = [str(v).strip() if v else "" for v in next(ws2.iter_rows(min_row=3, max_row=3, values_only=True))]
    rows2 = [list(row) for row in ws2.iter_rows(min_row=4, values_only=True)]
    parse_summary(rows2, hdr2)

    # 2026.xlsx 各拠点シート合計行: R7.9以降（調整ファイルから取得）
    SHEET_CODE = {
        "本社": "612-1", "公共": "612-2", "遠藤": "614-2",
        "板橋": "615-1", "新横浜": "615-2", "有価物": "615-4",
    }
    for sheet_name, code in SHEET_CODE.items():
        ws_i = wb2[sheet_name]
        all_rows = list(ws_i.iter_rows(values_only=True))

        # ヘッダー行を特定
        hdr_idx = None
        for ri, row in enumerate(all_rows):
            if any(v and str(v).strip().startswith("R") and "." in str(v) for v in row):
                hdr_idx = ri; break
        if hdr_idx is None: continue

        col_ym = parse_month_header(all_rows[hdr_idx])
        r77_col = next((ci for ci, ym in col_ym.items() if ym == "2025-07"), None)
        if r77_col is None: continue

        # R7.7の既知合計値で合計行を特定（ダブルカウント回避）
        anchor = sales.get("2025-07", {}).get(code, 0)
        if anchor == 0: continue

        total_row = None
        for row in all_rows:
            if r77_col < len(row) and row[r77_col]:
                try:
                    if abs(float(row[r77_col]) - anchor) < 10:
                        total_row = list(row); break
                except: pass
        if total_row is None: continue

        # R7.9以降の値を追加
        for ci, ym in col_ym.items():
            if ym <= "2025-08": continue  # 売上シートから取得済
            if ci >= len(total_row) or not total_row[ci]: continue
            try:
                v = float(total_row[ci])
                if v > 1000:  # ゴミ値除外
                    if ym not in sales:
                        sales[ym] = {}
                    sales[ym][code] = v
            except: pass

    # _total が未計算の月は合算
    for ym, d in sales.items():
        if "_total" not in d:
            d["_total"] = sum(v for k, v in d.items() if k != "_total")

    return {k: v for k, v in sorted(sales.items()) if any(v > 0 for v in v.values())}

# ════════════════════════════════════════════════════════════════
#  SHEET 1: 案件別採算
# ════════════════════════════════════════════════════════════════
def sheet_anken(ws, sales):
    ws.title = "案件別採算"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C4"

    # タイトル
    ws.merge_cells("A1:T1")
    t = ws["A1"]
    t.value = "案件別 採算管理　（★黄色セル＝入力　緑セル＝自動計算）"
    t.font = f(bold=True, sz=14, color=WHITE)
    t.fill = fill(NAVY); t.alignment = a(h="center"); ws.row_dimensions[1].height = 34

    # 説明
    ws.merge_cells("A2:T2")
    note = ws["A2"]
    note.value = "【売上】⑤売上原票シートをVLOOKUPで参照（D列=集計月 YYYY-MM を入力すると自動セット）　【原価】黄色セルに毎月入力してください"
    note.font = f(sz=9, color=GRAY); note.fill = fill(LGRAY); note.alignment = a(h="left"); note.border = b()
    ws.row_dimensions[2].height = 16

    # カラム幅
    widths = [10, 22, 11,  14, 14, 16,  14, 12, 12, 12, 12, 11, 11, 16,  14, 10, 12,  12, 16, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ヘッダー（2段）
    ws.row_dimensions[3].height = 16
    ws.row_dimensions[4].height = 38
    def merge_hd(r1,c1,r2,c2,txt,bg):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        cell = ws.cell(row=r1,column=c1,value=txt)
        cell.font=f(bold=True,sz=9,color=WHITE); cell.fill=fill(bg)
        cell.alignment=a(h="center",v="center",wrap=True); cell.border=b()

    merge_hd(3,1,4,1,"案件ID",    NAVY)
    merge_hd(3,2,4,2,"案件名",    NAVY)
    merge_hd(3,3,4,3,"集計月\n(YYYY-MM)", NAVY)
    # 売上
    merge_hd(3,4,3,6,"売上（VLOOKUPで自動参照）","155A00")
    for col, txt in [(4,"運送\n612-1"),(5,"入札\n612-2"),(6,"板橋\n615-1")]:
        merge_hd(4,col,4,col,txt,"155A00")
    merge_hd(3,7,4,7,"その他\n売上合計","155A00")
    merge_hd(3,8,4,8,"売上\n合計","155A00")
    # 原価
    merge_hd(3,9,3,15,"原価　（黄色セルに入力）","7B3F00")
    for col,txt in [(9,"人件費"),(10,"燃料費"),(11,"車両\n維持費"),(12,"中間\n処理費"),
                    (13,"搬入\n運搬費"),(14,"電子M\n代行費"),(15,"管理費")]:
        merge_hd(4,col,4,col,txt,"7B3F00")
    merge_hd(3,16,4,16,"原価\n合計","7B3F00")
    # 採算
    merge_hd(3,17,4,17,"粗利益","1A6B2E")
    merge_hd(3,18,4,18,"粗利率","1A6B2E")
    merge_hd(3,19,4,19,"採算\n判定","1A6B2E")
    merge_hd(3,20,4,20,"前月比\n売上","1A6B2E")

    JPY = "#,##0"
    PCT = "0.0%"
    months = sorted(sales.keys())

    ROW = 5
    prev_row = {}   # cid -> last row index
    for cid, info in CASE.items():
        bg = info["bg"]
        hd = info["hd"]
        name = info["name"]
        pct_col = {"C001":4,"C002":5,"C003":6}[cid]  # マッピング列

        for ym in months:
            ws.row_dimensions[ROW].height = 22

            c(ws,ROW,1,cid,   bold=True,sz=11,bg=bg,ha="center")
            c(ws,ROW,2,name,  sz=11,bg=bg)

            # D列: 集計月（入力キー）
            cx = ws.cell(row=ROW,column=3,value=ym)
            cx.font=f(bold=True,sz=11,color=WHITE); cx.fill=fill(hd)
            cx.alignment=a(h="center"); cx.border=b()

            # 売上（VLOOKUP）
            # E=612-1分, F=612-2分, G=615-1分
            for col, code, src_col in [(4,"612-1",3),(5,"612-2",4),(6,"615-1",10)]:
                vl = ws.cell(row=ROW,column=col)
                vl.value = (f'=IFERROR(VLOOKUP(C{ROW},▼売上原票!$A:$P,{src_col},0)'
                            f'*VLOOKUP("{code}",▼マッピング!$A:$G,{pct_col},0)/100,0)')
                vl.number_format=JPY; vl.font=f(sz=10)
                vl.fill=fill("E8F5E9"); vl.alignment=a(h="right"); vl.border=b()

            # G: その他科目合計
            other_codes = [("613-1",5),("613-2",6),("614-1",7),("614-2",8),("614-3",9),
                           ("615-2",11),("615-3",12),("615-4",13)]
            other_expr = "+".join(
                f'IFERROR(VLOOKUP(C{ROW},▼売上原票!$A:$P,{sc_},0)*VLOOKUP("{cd}",▼マッピング!$A:$G,{pct_col},0)/100,0)'
                for cd, sc_ in other_codes
            )
            g_ = ws.cell(row=ROW,column=7)
            g_.value = f"={other_expr}"
            g_.number_format=JPY; g_.font=f(sz=10)
            g_.fill=fill("E8F5E9"); g_.alignment=a(h="right"); g_.border=b()

            # H: 売上合計
            h_ = ws.cell(row=ROW,column=8)
            h_.value = f"=D{ROW}+E{ROW}+F{ROW}+G{ROW}"
            h_.number_format=JPY; h_.font=f(bold=True,sz=12)
            h_.fill=fill("C8F7D4"); h_.alignment=a(h="right"); h_.border=b()

            # 原価（黄色）空欄
            for ci in range(9, 16):
                c(ws,ROW,ci,None,sz=10,bg=YEL,ha="right",fmt=JPY)

            # P: 原価合計
            p_ = ws.cell(row=ROW,column=16)
            p_.value = f"=I{ROW}+J{ROW}+K{ROW}+L{ROW}+M{ROW}+N{ROW}+O{ROW}"
            p_.number_format=JPY; p_.font=f(bold=True,sz=11)
            p_.fill=fill(YEL); p_.alignment=a(h="right"); p_.border=b()

            # Q: 粗利
            q_ = ws.cell(row=ROW,column=17)
            q_.value = f"=H{ROW}-P{ROW}"
            q_.number_format=JPY; q_.font=f(bold=True,sz=12)
            q_.fill=fill(GRNLT); q_.alignment=a(h="right"); q_.border=b()

            # R: 粗利率
            r_ = ws.cell(row=ROW,column=18)
            r_.value = f"=IF(H{ROW}<>0,Q{ROW}/H{ROW},0)"
            r_.number_format=PCT; r_.font=f(bold=True,sz=12)
            r_.fill=fill(GRNLT); r_.alignment=a(h="center"); r_.border=b()

            # S: 採算判定
            s_ = ws.cell(row=ROW,column=19)
            s_.value = (f'=IF(R{ROW}>=0.25,"◎ 優良",'
                        f'IF(R{ROW}>=0.15,"○ 良好",'
                        f'IF(R{ROW}>=0.05,"△ 要注意","× 赤字")))')
            s_.font=f(bold=True,sz=11); s_.fill=fill(GRNLT)
            s_.alignment=a(h="center"); s_.border=b()

            # T: 前月比
            if cid in prev_row:
                t_ = ws.cell(row=ROW,column=20)
                t_.value = f"=IF(H{prev_row[cid]}<>0,H{ROW}/H{prev_row[cid]}-1,\"\")"
                t_.number_format="+0.0%;-0.0%;0.0%"; t_.font=f(sz=10)
                t_.fill=fill(bg); t_.alignment=a(h="center"); t_.border=b()
            else:
                c(ws,ROW,20,"—",sz=10,bg=bg,ha="center")

            prev_row[cid] = ROW
            ROW += 1

        # 案件区切り線（空行）
        ws.row_dimensions[ROW].height = 8
        for col in range(1,21):
            ws.cell(row=ROW,column=col).fill = fill(NAVY)
        ROW += 1

    # 入力用空白行（将来月）
    for _ in range(15):
        ws.row_dimensions[ROW].height = 22
        bg = LGRAY if ROW%2==0 else WHITE
        c(ws,ROW,1,"",sz=11,bg=bg,ha="center")
        c(ws,ROW,2,"",sz=11,bg=bg)
        c(ws,ROW,3,"",sz=11,bg=bg,ha="center")
        for col in range(4,8):
            cc = ws.cell(row=ROW,column=col)
            cc.value=""
            cc.number_format=JPY; cc.font=f(sz=10)
            cc.fill=fill("E8F5E9"); cc.alignment=a(h="right"); cc.border=b()
        for col in range(9,16):
            c(ws,ROW,col,None,sz=10,bg=YEL,ha="right",fmt=JPY)
        for col,expr,nm in [
            (8,  f"=D{ROW}+E{ROW}+F{ROW}+G{ROW}", JPY),
            (16, f"=IF(I{ROW}+J{ROW}+K{ROW}+L{ROW}+M{ROW}+N{ROW}+O{ROW}=0,\"\",I{ROW}+J{ROW}+K{ROW}+L{ROW}+M{ROW}+N{ROW}+O{ROW})", JPY),
            (17, f"=IF(H{ROW}<>0,H{ROW}-P{ROW},\"\")", JPY),
            (18, f"=IF(H{ROW}<>0,Q{ROW}/H{ROW},\"\")", PCT),
            (19, f'=IF(R{ROW}="","",IF(R{ROW}>=0.25,"◎ 優良",IF(R{ROW}>=0.15,"○ 良好",IF(R{ROW}>=0.05,"△ 要注意","× 赤字"))))', "@"),
        ]:
            cc = ws.cell(row=ROW,column=col)
            cc.value=expr; cc.number_format=nm
            cc.font=f(bold=(col in(8,16,17,18)),sz=11 if col in(8,17,18) else 10)
            cc.fill=fill("E8F5E9" if col in(4,5,6,7) else YEL if 9<=col<=15 else GRNLT if col>=16 else bg)
            cc.alignment=a(h="center" if col==19 else "right"); cc.border=b()
        ROW += 1

    # 条件付き書式
    end = ROW-1
    for expr, fg_c, bg_c in [
        ('S5="◎ 優良"', "145A32", "C8F7D4"),
        ('S5="○ 良好"', "1A6B2E", GRNLT),
        ('S5="△ 要注意"',"856404", YEL),
        ('S5="× 赤字"', "8B0000", REDLT),
    ]:
        ws.conditional_formatting.add(f"S5:S{end}",
            FormulaRule(formula=[expr], fill=fill(bg_c), font=Font(bold=True,color=fg_c)))
    ws.conditional_formatting.add(f"R5:R{end}",
        ColorScaleRule(start_type="num",start_value=0,   start_color="FF6B6B",
                       mid_type="num",  mid_value=0.15,  mid_color="FFD93D",
                       end_type="num",  end_value=0.30,  end_color="6BCB77"))


# ════════════════════════════════════════════════════════════════
#  SHEET 2: ルート別採算
# ════════════════════════════════════════════════════════════════
def sheet_route(ws, sales):
    ws.title = "ルート別採算"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D4"

    ws.merge_cells("A1:S1")
    t = ws["A1"]
    t.value = "ルート別 採算管理　（ルート = 科目コード / 各営業担当）"
    t.font=f(bold=True,sz=14,color=WHITE); t.fill=fill(NAVY)
    t.alignment=a(h="center"); ws.row_dimensions[1].height=34

    ws.merge_cells("A2:S2")
    note=ws["A2"]
    note.value = "【売上】⑤売上原票からVLOOKUPで自動参照　【原価】黄色セルに実績を入力　各ルートの粗利率で採算の良し悪しを確認できます"
    note.font=f(sz=9,color=GRAY); note.fill=fill(LGRAY); note.alignment=a(h="left"); note.border=b()
    ws.row_dimensions[2].height=16

    # カラム幅
    widths = [11,18,11, 16, 14,12,12,12,11,11,14, 14,10,12, 12,16,16]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    # ヘッダー
    ws.row_dimensions[3].height=16; ws.row_dimensions[4].height=36
    def mhd(r1,c1,r2,c2,txt,bg):
        ws.merge_cells(start_row=r1,start_column=c1,end_row=r2,end_column=c2)
        cell=ws.cell(row=r1,column=c1,value=txt)
        cell.font=f(bold=True,sz=9,color=WHITE); cell.fill=fill(bg)
        cell.alignment=a(h="center",v="center",wrap=True); cell.border=b()

    mhd(3,1,4,1,"ルートコード",NAVY); mhd(3,2,4,2,"ルート名",NAVY)
    mhd(3,3,4,3,"集計月\n(YYYY-MM)",NAVY)
    mhd(3,4,4,4,"売上\n(VLOOKUP)","155A00")
    mhd(3,5,3,11,"原価　（黄色セル入力）","7B3F00")
    for col,txt in [(5,"人件費"),(6,"燃料費"),(7,"車両費"),(8,"処理費"),
                    (9,"運搬費"),(10,"その他"),(11,"合計")]:
        mhd(4,col,4,col,txt,"7B3F00")
    mhd(3,12,4,12,"粗利益","1A6B2E")
    mhd(3,13,4,13,"粗利率","1A6B2E")
    mhd(3,14,4,14,"採算判定","1A6B2E")
    mhd(3,15,4,15,"主担当\n案件","1A6B2E")
    mhd(3,16,4,16,"前月比\n売上","1A6B2E")
    mhd(3,17,4,17,"月間回収量\n(kg推算)",GRAY)

    JPY="#,##0"; PCT="0.0%"

    # ルートの色グループ
    route_bg = {
        "612-1":BLULT,"612-2":BLULT,    # 本社系
        "613-1":PURLT,"613-2":PURLT,    # 車輌系
        "614-1":GRNLT,"614-2":GRNLT,"614-3":GRNLT,  # 現場系
        "615-1":BLULT,"615-2":BLULT,    # 営業所系
        "615-3":"FFF9C4","615-4":GRNLT, # 現金・有価物
    }
    # 主担当案件（マッピング設定から）
    route_case = {}
    for code, (p1,p2,p3) in MAPPING.items():
        if p1>=p2 and p1>=p3 and p1>0: route_case[code]="C001東急"
        elif p2>=p1 and p2>=p3 and p2>0: route_case[code]="C002都営"
        elif p3>0: route_case[code]="C003一般"
        else: route_case[code]="別管理"

    # 原価の目安（月次参考値）
    cost_guide = {
        "612-1":(650000,190000,85000,0,75000,45000),
        "612-2":(1800000,520000,240000,1200000,180000,120000),
        "613-1":(180000,55000,180000,0,15000,20000),
        "613-2":(0,0,0,0,0,0),
        "614-1":(85000,25000,15000,0,20000,10000),
        "614-2":(120000,38000,18000,0,25000,12000),
        "614-3":(20000,6000,3000,0,5000,3000),
        "615-1":(850000,248000,118000,580000,95000,62000),
        "615-2":(280000,82000,38000,195000,32000,20000),
        "615-3":(15000,4000,2000,0,3000,2000),
        "615-4":(35000,10000,5000,0,8000,5000),
    }

    months = sorted(sales.keys())
    ROW = 5
    prev_row = {}

    for code, rname, _ in ROUTES:
        bg = route_bg.get(code, LGRAY)
        src_col = 3 + CODES.index(code)   # ⑤売上原票の列番号

        for ym in months:
            ws.row_dimensions[ROW].height=21

            c(ws,ROW,1,code,bold=True,sz=11,bg=bg,ha="center")
            c(ws,ROW,2,rname,sz=11,bg=bg)

            cx=ws.cell(row=ROW,column=3,value=ym)
            cx.font=f(bold=True,sz=10,color=WHITE); cx.fill=fill(NAVY)
            cx.alignment=a(h="center"); cx.border=b()

            # 売上（VLOOKUP from ▼売上原票）
            vl=ws.cell(row=ROW,column=4)
            vl.value=f"=IFERROR(VLOOKUP(C{ROW},▼売上原票!$A:$P,{src_col},0),0)"
            vl.number_format=JPY; vl.font=f(bold=True,sz=11)
            vl.fill=fill("E8F5E9"); vl.alignment=a(h="right"); vl.border=b()

            # 原価（黄色入力欄）空欄
            for ci in range(5, 11):
                c(ws,ROW,ci,None,sz=10,bg=YEL,ha="right",fmt=JPY)

            # 原価合計
            k_=ws.cell(row=ROW,column=11)
            k_.value=f"=E{ROW}+F{ROW}+G{ROW}+H{ROW}+I{ROW}+J{ROW}"
            k_.number_format=JPY; k_.font=f(bold=True,sz=11)
            k_.fill=fill(YEL); k_.alignment=a(h="right"); k_.border=b()

            # 粗利
            l_=ws.cell(row=ROW,column=12)
            l_.value=f"=D{ROW}-K{ROW}"
            l_.number_format=JPY; l_.font=f(bold=True,sz=12)
            l_.fill=fill(GRNLT); l_.alignment=a(h="right"); l_.border=b()

            # 粗利率
            m_=ws.cell(row=ROW,column=13)
            m_.value=f"=IF(D{ROW}<>0,L{ROW}/D{ROW},0)"
            m_.number_format=PCT; m_.font=f(bold=True,sz=12)
            m_.fill=fill(GRNLT); m_.alignment=a(h="center"); m_.border=b()

            # 採算判定
            n_=ws.cell(row=ROW,column=14)
            n_.value=(f'=IF(D{ROW}=0,"—",'
                      f'IF(M{ROW}>=0.25,"◎ 優良",'
                      f'IF(M{ROW}>=0.15,"○ 良好",'
                      f'IF(M{ROW}>=0.05,"△ 要注意","× 赤字"))))')
            n_.font=f(bold=True,sz=10); n_.fill=fill(GRNLT)
            n_.alignment=a(h="center"); n_.border=b()

            # 主担当案件
            c(ws,ROW,15,route_case.get(code,"—"),sz=10,bg=bg,ha="center")

            # 前月比
            if code in prev_row:
                o_=ws.cell(row=ROW,column=16)
                o_.value=f"=IF(D{prev_row[code]}<>0,D{ROW}/D{prev_row[code]}-1,\"\")"
                o_.number_format="+0.0%;-0.0%;0.0%"; o_.font=f(sz=10)
                o_.fill=fill(bg); o_.alignment=a(h="center"); o_.border=b()
            else:
                c(ws,ROW,16,"—",sz=10,bg=bg,ha="center")
            prev_row[code]=ROW

            # kg推算（売上÷想定単価100円/kgで概算）
            p_=ws.cell(row=ROW,column=17)
            p_.value=f"=IFERROR(D{ROW}/100,0)"
            p_.number_format="#,##0"; p_.font=f(sz=10,color=GRAY)
            p_.fill=fill(bg); p_.alignment=a(h="right"); p_.border=b()

            ROW+=1

        # ルート区切り
        ws.row_dimensions[ROW].height=5
        for col in range(1,18):
            ws.cell(row=ROW,column=col).fill=fill("CCDDEE")
        ROW+=1

    # 条件付き書式
    end=ROW-1
    for expr,fg_c,bg_c in [
        ('N5="◎ 優良"',"145A32","C8F7D4"),
        ('N5="○ 良好"',"1A6B2E",GRNLT),
        ('N5="△ 要注意"',"856404",YEL),
        ('N5="× 赤字"',"8B0000",REDLT),
    ]:
        ws.conditional_formatting.add(f"N5:N{end}",
            FormulaRule(formula=[expr],fill=fill(bg_c),font=Font(bold=True,color=fg_c)))
    ws.conditional_formatting.add(f"M5:M{end}",
        ColorScaleRule(start_type="num",start_value=0,  start_color="FF6B6B",
                       mid_type="num", mid_value=0.15,  mid_color="FFD93D",
                       end_type="num", end_value=0.30,  end_color="6BCB77"))


# ════════════════════════════════════════════════════════════════
#  SHEET 3: ▼売上原票（VLOOKUP元データ）
# ════════════════════════════════════════════════════════════════
def sheet_uriagehyo(ws, sales):
    ws.title = "▼売上原票"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    ws.merge_cells("A1:P1")
    t=ws["A1"]
    t.value="▼売上原票　月次売上実績（科目別・税抜）　※このシートは直接編集しないでください"
    t.font=f(bold=True,sz=13,color=WHITE); t.fill=fill("335577")
    t.alignment=a(h="center"); ws.row_dimensions[1].height=30

    ws.merge_cells("A2:P2")
    n2=ws["A2"]
    n2.value="【毎月の更新方法】 新しい月のデータを最終行の下に1行追加してください。A列=YYYY-MM、C〜M列=各科目売上（税抜円）"
    n2.font=f(sz=9,color=GRAY); n2.fill=fill(LGRAY); n2.alignment=a(h="left"); n2.border=b()
    ws.row_dimensions[2].height=16

    widths=[13,9]+[13]*11+[15,11,15]
    for i,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    ws.row_dimensions[3].height=36
    for col,(lbl,w) in enumerate([
        ("年月\n★VLOOKUPキー",13),("R表記",9),
        *[(f"{cd}\n{nm}",13) for cd,nm,_ in ROUTES],
        ("合計\n(税抜)",15),("消費税\n(△)",11),("合計\n(税込)",15)
    ],1):
        c(ws,3,col,lbl,bold=True,sz=9,bg=BLUE,fc=WHITE,ha="center",wrap=True)

    months = sorted(sales.keys())
    for r, ym in enumerate(months, 4):
        ws.row_dimensions[r].height=19
        d = sales[ym]
        era=int(ym[:4])-2018
        era_str=f"R{era}.{int(ym[5:7])}"
        total=d.get("_total",0)
        tax=-round(total*10/11)
        bg=LGRAY if r%2==0 else WHITE

        c(ws,r,1,ym,bold=True,sz=10,bg=bg,ha="center")
        c(ws,r,2,era_str,sz=10,bg=bg,ha="center")
        for ci,code in enumerate(CODES,3):
            v=d.get(code,0)
            c(ws,r,ci,v if v else None,sz=10,bg=bg,ha="right",fmt="#,##0")
        c(ws,r,14,total if total else None,bold=True,sz=11,bg=bg,ha="right",fmt="#,##0")
        c(ws,r,15,tax if total else None,sz=10,bg=bg,ha="right",fmt="#,##0")
        cell16=ws.cell(row=r,column=16)
        cell16.value=f"=IF(N{r}<>\"\",N{r}+O{r},\"\")"
        cell16.number_format="#,##0"; cell16.font=f(bold=True,sz=10)
        cell16.fill=fill(bg); cell16.alignment=a(h="right"); cell16.border=b()

    # 将来月入力枠
    for i in range(12):
        r=3+len(months)+1+i
        ws.row_dimensions[r].height=19
        bg=LGRAY if r%2==0 else WHITE
        c(ws,r,1,"←ここに入力",sz=9,bg="FFFDE7",ha="center",fc="856404")
        c(ws,r,2,"",sz=10,bg=bg,ha="center")
        for ci in range(3,14):
            c(ws,r,ci,None,sz=10,bg=bg,ha="right",fmt="#,##0")
        auto=f"=IF(C{r}+D{r}+E{r}+F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r}+M{r}=0,\"\",C{r}+D{r}+E{r}+F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r}+M{r})"
        c(ws,r,14,auto,bold=True,sz=10,bg=bg,ha="right",fmt="#,##0")
        cell15=ws.cell(row=r,column=15)
        cell15.value=f"=IF(N{r}=\"\",\"\",ROUND(-N{r}/11,0))"
        cell15.number_format="#,##0"; cell15.font=f(sz=10)
        cell15.fill=fill(bg); cell15.alignment=a(h="right"); cell15.border=b()
        cell16=ws.cell(row=r,column=16)
        cell16.value=f"=IF(N{r}=\"\",\"\",N{r}+O{r})"
        cell16.number_format="#,##0"; cell16.font=f(bold=True,sz=10)
        cell16.fill=fill(bg); cell16.alignment=a(h="right"); cell16.border=b()


# ════════════════════════════════════════════════════════════════
#  SHEET 4: ▼マッピング設定
# ════════════════════════════════════════════════════════════════
def sheet_mapping(ws):
    ws.title = "▼マッピング設定"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:G1")
    t=ws["A1"]
    t.value="▼マッピング設定　ルート(科目)→案件の配分率　★初回のみ確認・調整してください"
    t.font=f(bold=True,sz=13,color=WHITE); t.fill=fill("335577")
    t.alignment=a(h="center"); ws.row_dimensions[1].height=30

    ws.merge_cells("A2:G2")
    n2=ws["A2"]
    n2.value="各行の D+E+F=100 になるよう設定してください。615-1(板橋)は東急グループと一般事業者の両方を担当しているため分割設定しています。"
    n2.font=f(sz=9,color=GRAY); n2.fill=fill(LGRAY); n2.alignment=a(h="left"); n2.border=b()
    ws.row_dimensions[2].height=16

    for i,w in enumerate([12,18,30,14,14,14,14],1):
        ws.column_dimensions[get_column_letter(i)].width=w

    ws.row_dimensions[3].height=36
    for col,lbl in enumerate(["科目コード","科目名","配分根拠（メモ）",
                               "C001\n東急グループ%","C002\n都営地下鉄%",
                               "C003\n一般事業者%","合計%\n(100が正常)"],1):
        c(ws,3,col,lbl,bold=True,sz=9,bg=NAVY,fc=WHITE,ha="center",wrap=True)

    reasons = {
        "612-1":"本社が受注する民間運送 → 一般事業者に100%",
        "612-2":"行政・公共入札案件 → 都営地下鉄・区役所等に100%",
        "613-1":"車輌売上（全案件に按分）",
        "613-2":"KCサキュレ分（別管理・0%）",
        "614-1":"コンテナ → 一般事業者に100%",
        "614-2":"ENDO → 一般事業者に100%",
        "614-3":"現金取引 → 一般事業者に100%",
        "615-1":"板橋営業所：東急グループ拠点60%・一般40% ★実績に応じて調整",
        "615-2":"新横浜営業所 → 一般事業者に100%",
        "615-3":"現金2 → 一般事業者に100%",
        "615-4":"有価物 → 一般事業者に100%",
    }

    for r, (code, rname, _) in enumerate(ROUTES, 4):
        ws.row_dimensions[r].height=28
        bg=LGRAY if r%2==0 else WHITE
        p1,p2,p3=MAPPING[code]

        c(ws,r,1,code, bold=True,sz=11,bg=bg,ha="center")
        c(ws,r,2,rname,sz=11,bg=bg)
        c(ws,r,3,reasons.get(code,""),sz=9,bg=bg,wrap=True,fc=GRAY)

        # 入力セル（配分率）
        for col,val,bg_c in [(4,p1,"E8F5E9"),(5,p2,"F3E5F5"),(6,p3,"E0F2F1")]:
            cell=ws.cell(row=r,column=col,value=val)
            cell.number_format="0"; cell.font=f(bold=True,sz=13)
            cell.fill=fill(bg_c); cell.alignment=a(h="center"); cell.border=b()

        # 合計確認
        g_=ws.cell(row=r,column=7)
        g_.value=f"=D{r}+E{r}+F{r}"
        g_.number_format="0"
        g_.font=Font(bold=True,size=13,name="Yu Gothic UI")
        g_.fill=fill("C8F7D4"); g_.alignment=a(h="center"); g_.border=b()
        ws.conditional_formatting.add(f"G{r}:G{r}",
            FormulaRule(formula=[f"G{r}<>100"],
                        fill=fill(REDLT),font=Font(bold=True,color="8B0000")))

    # 凡例
    nr=4+len(ROUTES)+1
    ws.merge_cells(f"A{nr}:G{nr}")
    legend=ws.cell(row=nr,column=1,
        value="【判定基準】 ◎ 優良=粗利率25%以上　○ 良好=15〜25%　△ 要注意=5〜15%　× 赤字=5%未満")
    legend.font=f(sz=10,color=NAVY); legend.fill=fill(LGRAY)
    legend.alignment=a(h="center"); legend.border=b()
    ws.row_dimensions[nr].height=22


# ════════════════════════════════════════════════════════════════
#  SHEET 5: 従業員マスター
# ════════════════════════════════════════════════════════════════
# (所属, 氏名, 区分, 金額, 直間, 備考)
# 区分: 月給=固定月額, 日当=日当額×稼働日数, 時給=時給×稼働時間, 自転車=日当8447×稼働日数
SALARY_DATA = [
    # ── 本社 月給制 ──────────────────────────────────
    ("本社","横井 哲也",  "月給",269545,"直接",""),
    ("本社","上林 幸二",  "月給",254545,"直接",""),
    ("本社","加藤 光裕",  "月給",283545,"直接",""),
    ("本社","吉田 創",    "月給",249395,"直接",""),
    ("本社","原口 陽一",  "月給",243155,"直接",""),
    ("本社","北川 善己",  "月給",266955,"直接","東京発送手当"),
    ("本社","平野 裕幸",  "月給",263155,"直接","渡邊ベニヤ"),
    ("本社","山口 晃司",  "月給",243155,"直接","いずみ"),
    ("本社","石垣眞一郎", "月給",263155,"直接","菱倉"),
    ("本社","高島 正人",  "月給",246155,"直接","いずみ"),
    ("本社","綱島 博隆",  "月給",268155,"直接","千代田区"),
    ("本社","後藤 実",    "月給",243155,"直接",""),
    ("本社","嶋村 真実",  "月給",243155,"直接",""),
    ("本社","岡田 学",    "月給",243155,"間接","文書"),
    ("本社","横山 信廣",  "月給",243155,"直接",""),
    ("本社","笹沼 信人",  "月給",263155,"直接","東京発送手当"),
    ("本社","諏訪 清行",  "月給",243155,"直接",""),
    ("本社","清水 孝之",  "月給",246155,"直接","いずみ"),
    ("本社","山下 勝二",  "月給",243155,"直接",""),
    ("本社","根本 純一",  "月給",278155,"直接","自転車撤去"),
    ("本社","花木 凌",    "月給",243155,"直接",""),
    ("本社","川井 良之",  "月給",243155,"直接",""),
    # ── 本社 日当制 ──────────────────────────────────
    ("本社","岡村 栄",    "日当",10000,"直接",""),
    ("本社","川端 保",    "日当",11000,"直接",""),
    ("本社","鈴木 賢二",  "日当",10000,"直接",""),
    ("本社","小西 正悦",  "日当",10000,"直接",""),
    ("本社","西開地良二", "日当",10000,"直接","いずみ手当"),
    ("本社","四位 宗光",  "日当", 9320,"直接",""),
    ("本社","赤崎 博隆",  "日当",10000,"直接",""),
    ("本社","近藤 政夫",  "日当",10000,"直接",""),
    ("本社","佐藤 淳司",  "日当",10000,"直接","自転車撤去"),
    ("本社","門脇 明敏",  "日当",10000,"直接","いずみ"),
    ("本社","小島 一彦",  "日当",10000,"直接",""),
    ("本社","後藤 保幸",  "日当",10000,"直接",""),
    ("本社","昆野 良博",  "日当",10000,"直接",""),
    ("本社","伊藤 敏正",  "日当",10000,"直接",""),
    ("本社","石井 努",    "日当",10000,"直接","いずみ"),
    ("本社","金川 裕",    "日当",10000,"直接","文書交換"),
    ("本社","大澤 弘直",  "日当", 8000,"直接","6.5h"),
    ("本社","羽田 美惠",  "日当",10000,"間接",""),
    ("本社","赤石 政春",  "日当",10000,"直接","文書交換"),
    ("本社","小山内文夫", "日当",10000,"直接","千代田区"),
    ("本社","林 重晴",    "日当",10000,"直接","いずみ"),
    ("本社","中原 健一",  "日当",10000,"直接","文書交換"),
    ("本社","樋口 隆一",  "日当",10000,"直接","自転車撤去"),
    ("本社","中村 史德",  "日当",10000,"直接","自転車撤去"),
    ("本社","金子 豊",    "日当",10000,"直接","自転車撤去"),
    ("本社","斉藤 善",    "日当",10000,"直接","自転車撤去"),
    ("本社","松崎 亨",    "日当",10000,"直接",""),
    ("本社","町田 昌明",  "日当",10000,"直接",""),
    ("本社","堀 富雄",    "日当",10000,"直接",""),
    ("本社","栗原文次郎", "日当",10000,"直接",""),
    ("本社","高栁 恵州",  "日当",10000,"直接",""),
    ("本社","吉見 之男",  "日当",10000,"直接",""),
    ("本社","坂本 龍男",  "日当",10000,"直接",""),
    ("本社","桃谷 政良",  "日当",10000,"直接",""),
    ("本社","竹内 幸一",  "日当",10000,"直接",""),
    ("本社","林 幸夫",    "日当",10000,"直接",""),
    # ── 本社 時給制 ──────────────────────────────────
    ("本社","加藤実都代", "時給",1805,"間接",""),
    ("本社","林 典子",    "時給",1165,"間接",""),
    ("本社","足立秀次郎", "時給",1165,"間接","文書交換室"),
    ("本社","松原 光年",  "時給",1165,"間接","文書交換室"),
    ("本社","吉川 洋平",  "時給",1165,"間接","文書交換室"),
    ("本社","野村 正信",  "時給",1165,"間接","文書交換室"),
    ("本社","野原 典子",  "時給",1165,"間接","文書交換室"),
    ("本社","坂本多枝子", "時給",1165,"間接","文書交換室"),
    ("本社","矢作 明美",  "時給",1165,"間接","文書交換室"),
    ("本社","木村 貴子",  "時給",1165,"間接","文書交換室"),
    ("本社","稲富 笑子",  "時給",1165,"間接","文書交換室"),
    ("本社","芳本 智代",  "時給",1165,"間接","文書交換室"),
    # ── 本社 自転車撤去助手 ───────────────────────────
    ("本社","村上 博美",  "自転車",8447,"直接","7.25h"),
    ("本社","渡邊 茂夫",  "自転車",8447,"直接","7.25h"),
    ("本社","川満 俊一",  "自転車",8447,"直接","7.25h"),
    ("本社","阿久津信介", "自転車",8447,"直接","7.25h 班長手当"),
    ("本社","竹内 和夫",  "自転車",8447,"直接","7.25h"),
    ("本社","橋本富士夫", "自転車",8447,"直接","7.25h"),
    ("本社","三浦 勲",    "自転車",8447,"直接","7.25h"),
    ("本社","青木 茂",    "自転車",8447,"直接","7.25h"),
    ("本社","後藤 高広",  "自転車",8447,"直接","7.25h"),
    ("本社","平山 進",    "自転車",8447,"直接","7.25h"),
    ("本社","小松 真裕",  "自転車",8447,"直接","7.25h"),
    ("本社","奥園 達雄",  "自転車",8447,"直接","7.25h"),
    ("本社","伊藤 直道",  "自転車",8447,"直接","7.25h"),
    ("本社","仲眞 栄治",  "自転車",8447,"直接","7.25h"),
    ("本社","白井 久雄",  "自転車",8447,"直接","7.25h"),
    ("本社","仲 宏樹",    "自転車",8447,"直接","7.25h 能率2"),
    ("本社","石井 英明",  "自転車",8447,"直接","7.25h"),
    ("本社","赤間 茂",    "自転車",8447,"直接","7.25h"),
    ("本社","宇田川博之", "自転車",8447,"直接","7.25h"),
    ("本社","菊池 義一",  "自転車",8447,"直接","7.25h"),
    ("本社","森 涼太",    "自転車",8447,"直接","7.25h 能率2"),
    # ── 板橋 月給制 ──────────────────────────────────
    ("板橋","舟山 真純",  "月給",278155,"直接","KC"),
    ("板橋","真栄城玄昌", "月給",278545,"直接",""),
    ("板橋","長野 義昭",  "月給",269395,"直接","深夜"),
    ("板橋","大森 茂",    "月給",276395,"直接","深夜"),
    ("板橋","柳沼 清志",  "月給",276395,"直接",""),
    ("板橋","内田 温城",  "月給",253155,"直接",""),
    ("板橋","大迫 二郎",  "月給",253155,"直接","深夜"),
    ("板橋","黒澤栄次郎", "月給",253155,"直接","深夜"),
    ("板橋","千葉 実",    "月給",253155,"直接","深夜"),
    ("板橋","宮原 茂",    "月給",253155,"直接","深夜"),
    ("板橋","堀 徳源",    "月給",263155,"直接",""),
    ("板橋","青木 心一",  "月給",333155,"直接",""),
    ("板橋","長谷川岳秀", "月給",253155,"直接","深夜"),
    ("板橋","小林 司",    "月給",253155,"直接","深夜"),
    ("板橋","中條 順一",  "月給",253155,"直接","深夜"),
    # ── 板橋 日当・時給 ──────────────────────────────
    ("板橋","吉橋 清",    "日当",10000,"直接","深夜手当"),
    ("板橋","志村 貴子",  "時給", 1165,"間接",""),
    ("板橋","加藤 裕子",  "時給", 1165,"間接",""),
    # ── 新横浜 月給制 ────────────────────────────────
    ("新横浜","富田 定夫","月給",248155,"直接",""),
    ("新横浜","細野 篤史","月給",293155,"直接","サッポロ"),
    ("新横浜","麻生 和義","月給",263155,"直接","サッポロ"),
    # ── 新横浜 日当・時給 ────────────────────────────
    ("新横浜","木幡 信行","日当",10000,"直接",""),
    ("新横浜","渡辺 義則","日当",10000,"直接",""),
    ("新横浜","府川 正明","日当",10000,"直接",""),
    ("新横浜","松尾 雅史","日当",10000,"直接","サッポロ"),
    ("新横浜","古川 美光","日当",10000,"直接","深夜"),
    ("新横浜","宮崎 守",  "日当",10000,"直接",""),
    ("新横浜","大矢 詔子","時給", 1165,"間接",""),
    # ── 埼玉 月給制 ──────────────────────────────────
    ("埼玉","和田 康弘",  "月給",263155,"直接",""),
    ("埼玉","藤内 順平",  "月給",253155,"直接","文京"),
    ("埼玉","土屋 紀元",  "月給",263155,"直接",""),
    ("埼玉","玉山 賢雄",  "月給",243155,"直接",""),
    ("埼玉","石橋 雅隆",  "月給",243155,"直接","ユニマテック"),
    ("埼玉","佐々木幸治", "月給",263155,"直接",""),
    ("埼玉","小堀 智由",  "月給",253155,"直接","文京"),
    ("埼玉","大森 翔太",  "月給",243155,"直接","ユニマテック"),
    ("埼玉","根岸 三鶴",  "月給",243155,"直接",""),
    ("埼玉","平澤 隼",    "月給",253155,"直接","文京"),
    ("埼玉","八峠 知之",  "月給",243155,"直接",""),
    ("埼玉","山田 和宏",  "月給",253155,"直接","文京"),
    ("埼玉","渡辺 貴之",  "月給",251155,"直接",""),
    ("埼玉","大嶋 隆",    "月給",243155,"直接",""),
    ("埼玉","日吉 眞次",  "月給",243155,"直接",""),
    ("埼玉","西川 一美",  "月給",253155,"直接","文京"),
    ("埼玉","篠原 幸一",  "月給",253155,"直接","文京"),
    ("埼玉","前崎 厚志",  "月給",253155,"直接","文京"),
    ("埼玉","小笠原良孝", "月給",243155,"直接",""),
    ("埼玉","金子 修治",  "月給",243155,"直接",""),
    ("埼玉","新海 善光",  "月給",258788,"直接","千代田区"),
    ("埼玉","北野 利春",  "月給",279288,"直接","千代田区"),
    ("埼玉","近藤 彰二",  "月給",272898,"直接","千代田区"),
    ("埼玉","和田 盛樹",  "月給",271898,"直接","千代田"),
    ("埼玉","折笠 良一",  "月給",272898,"直接","千代田区"),
    ("埼玉","山田 俊和",  "月給",266898,"直接","文京"),
    ("埼玉","荻野 泰宏",  "月給",266898,"直接","文京"),
    ("埼玉","寺木 達夫",  "月給",266898,"直接","文京"),
    ("埼玉","矢口 光春",  "月給",266898,"直接","文京"),
    # ── 埼玉 日当制 ──────────────────────────────────
    ("埼玉","宇都宮一房", "日当",10000,"直接",""),
    ("埼玉","大友 正昭",  "日当",10000,"直接","ユニマテック"),
    ("埼玉","高井 孝行",  "日当", 9000,"直接",""),
    ("埼玉","水野 敦夫",  "日当", 9000,"直接",""),
    ("埼玉","岩波 孝",    "日当",10000,"直接",""),
    ("埼玉","河野 大地",  "日当",10000,"直接",""),
    ("埼玉","本田 智彦",  "日当", 7500,"直接","点呼"),
    ("埼玉","磐城 孝司",  "日当", 7500,"直接","点呼"),
    ("埼玉","石塚 正海",  "日当", 7500,"直接","点呼"),
    ("埼玉","佐藤 芳治",  "日当", 7500,"直接",""),
    # ── 管理職 ───────────────────────────────────────
    ("管理","宇都宮 寛",  "月給",537000,"間接",""),
    ("管理","遠藤 一栄",  "月給",337000,"間接",""),
    ("管理","足立 孝子",  "月給",257993,"間接",""),
    ("管理","滝田 豪",    "月給",307830,"間接",""),
    ("管理","加藤 千明",  "月給",269601,"間接",""),
    ("管理","山口 博文",  "月給",250500,"間接",""),
    ("管理","佐藤 正美",  "月給",264500,"間接",""),
    ("管理","戸塚 秀喜",  "月給",260600,"間接",""),
    ("管理","落合 秋夫",  "月給",306298,"間接",""),
]

def sheet_employee_master(ws):
    ws.title = "▼従業員マスター"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C6"

    # タイトル
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = "従業員マスター（R6.10 新体系基準）　※給与データ機密 取扱注意"
    t.font = f(bold=True, sz=13, color=WHITE)
    t.fill = fill(NAVY); t.alignment = a(h="center"); ws.row_dimensions[1].height = 32

    # 注記
    ws.merge_cells("A2:I2")
    n = ws["A2"]
    n.value = "出典：令和6年10月変更 新体系　│　月給制=合計欄の固定額　│　日当制・時給制は稼働日数に応じて変動（E列入力）"
    n.font = f(sz=9, color=GRAY); n.fill = fill(LGRAY); n.alignment = a(h="left"); n.border = b()
    ws.row_dimensions[2].height = 16

    # 稼働日数設定
    ws.merge_cells("A3:C3")
    lbl = ws["A3"]
    lbl.value = "▼ 月間稼働日数（日当制・時給制の計算基準）"
    lbl.font = f(sz=10, color=NAVY, bold=True); lbl.fill = fill(LGRAY); lbl.border = b()
    wd_cell = ws["D3"]
    wd_cell.value = 22
    wd_cell.font = f(bold=True, sz=14, color=NAVY)
    wd_cell.fill = fill(YEL); wd_cell.alignment = a(h="center"); wd_cell.border = b()
    ws.merge_cells("E3:I3")
    hint = ws["E3"]
    hint.value = "← ここを変更すると日当制・時給制の月額が自動更新されます（時給制は×8時間/日で計算）"
    hint.font = f(sz=9, color=GRAY); hint.fill = fill(LGRAY); hint.border = b()
    ws.row_dimensions[3].height = 20

    ws.row_dimensions[4].height = 6  # スペーサー

    # ヘッダー
    ws.row_dimensions[5].height = 34
    for col, (txt, w) in enumerate([
        ("所属",8), ("氏名",14), ("雇用区分",10),
        ("月額/日当/時給",16), ("月間稼働日数\n(入力可)",14),
        ("月額人件費\n(自動計算)",16), ("直間\n区分",8), ("備考",24), ("参考:年間",14)
    ], 1):
        ws.column_dimensions[get_column_letter(col)].width = w
        cell = ws.cell(row=5, column=col, value=txt)
        cell.font = f(bold=True, sz=9, color=WHITE)
        cell.fill = fill(NAVY); cell.alignment = a(h="center", v="center", wrap=True); cell.border = b()

    JPY = "#,##0"
    LOCS = ["本社", "板橋", "新横浜", "埼玉", "管理"]
    LOC_COLORS = {"本社": BLULT, "板橋": "FFE0B2", "新横浜": "E8F5E9",
                  "埼玉": PURLT, "管理": "FAFAFA"}
    LOC_HD_COLORS = {"本社": "1565A8", "板橋": "E65100", "新横浜": "2E7D32",
                     "埼玉": "4A148C", "管理": "37474F"}
    TYPE_COLORS = {"月給": WHITE, "日当": YEL, "時給": YEL, "自転車": "FFF9C4"}

    ROW = 6
    current_loc = None
    loc_start = {}

    for (loc, name, typ, rate, role, note) in SALARY_DATA:
        # 拠点区切りヘッダー
        if loc != current_loc:
            if current_loc is not None:
                # 前拠点の小計行
                st = loc_start[current_loc]
                ws.row_dimensions[ROW].height = 18
                ws.merge_cells(f"A{ROW}:E{ROW}")
                stlbl = ws.cell(row=ROW, column=1,
                    value=f"{current_loc} 合計  ({ROW - st}名)")
                stlbl.font = f(bold=True, sz=10, color=WHITE)
                stlbl.fill = fill(LOC_HD_COLORS[current_loc])
                stlbl.alignment = a(h="right"); stlbl.border = b()
                tot = ws.cell(row=ROW, column=6,
                    value=f"=SUMIF(A{st}:A{ROW-1},\"{current_loc}\",F{st}:F{ROW-1})")
                tot.number_format = JPY; tot.font = f(bold=True, sz=12)
                tot.fill = fill(LOC_HD_COLORS[current_loc]); tot.alignment = a(h="right"); tot.border = b()
                for col in [7,8,9]:
                    ws.cell(row=ROW, column=col).fill = fill(LOC_HD_COLORS[current_loc])
                    ws.cell(row=ROW, column=col).border = b()
                ROW += 1

            # 新拠点ヘッダー
            loc_start[loc] = ROW
            ws.row_dimensions[ROW].height = 20
            ws.merge_cells(f"A{ROW}:I{ROW}")
            hd = ws.cell(row=ROW, column=1, value=f"■ {loc}")
            hd.font = f(bold=True, sz=12, color=WHITE)
            hd.fill = fill(LOC_HD_COLORS[loc])
            hd.alignment = a(h="left"); hd.border = b()
            ROW += 1
            current_loc = loc

        ws.row_dimensions[ROW].height = 20
        bg = LOC_COLORS[loc]
        tbg = TYPE_COLORS.get(typ, WHITE)

        c(ws, ROW, 1, loc,  sz=9, bg=bg, ha="center")
        c(ws, ROW, 2, name, sz=10, bg=bg)
        c(ws, ROW, 3, typ,  sz=9, bg=tbg, ha="center")

        # D: 金額
        d_ = ws.cell(row=ROW, column=4, value=rate)
        d_.number_format = JPY; d_.font = f(sz=10)
        d_.fill = fill(tbg); d_.alignment = a(h="right"); d_.border = b()

        # E: 稼働日数（月給は"-", その他は参照）
        e_ = ws.cell(row=ROW, column=5)
        if typ == "月給":
            e_.value = "—"
            e_.font = f(sz=10, color=GRAY); e_.fill = fill(bg)
            e_.alignment = a(h="center"); e_.border = b()
        else:
            e_.value = f"=$D$3"  # 共通稼働日数を参照
            e_.number_format = "0"; e_.font = f(sz=10)
            e_.fill = fill(YEL); e_.alignment = a(h="center"); e_.border = b()

        # F: 月額人件費
        f_ = ws.cell(row=ROW, column=6)
        if typ == "月給":
            f_.value = rate
        elif typ == "時給":
            f_.value = f"=D{ROW}*$D$3*8"   # 時給×稼働日数×8h
        else:  # 日当 or 自転車
            f_.value = f"=D{ROW}*$D$3"
        f_.number_format = JPY; f_.font = f(bold=True, sz=10)
        f_.fill = fill("D4EDDA"); f_.alignment = a(h="right"); f_.border = b()

        # G: 直間
        role_bg = "E8F5E9" if role == "直接" else REDLT
        c(ws, ROW, 7, role, sz=9, bg=role_bg, ha="center")

        # H: 備考
        c(ws, ROW, 8, note, sz=9, bg=bg, fc=GRAY)

        # I: 年間参考
        i_ = ws.cell(row=ROW, column=9)
        if typ == "月給":
            i_.value = rate * 12
        else:
            i_.value = f"=F{ROW}*12"
        i_.number_format = JPY; i_.font = f(sz=9, color=GRAY)
        i_.fill = fill(LGRAY); i_.alignment = a(h="right"); i_.border = b()

        ROW += 1

    # 最後の拠点小計
    if current_loc:
        st = loc_start[current_loc]
        ws.row_dimensions[ROW].height = 18
        ws.merge_cells(f"A{ROW}:E{ROW}")
        stlbl = ws.cell(row=ROW, column=1,
            value=f"{current_loc} 合計  ({ROW - st}名)")
        stlbl.font = f(bold=True, sz=10, color=WHITE)
        stlbl.fill = fill(LOC_HD_COLORS[current_loc])
        stlbl.alignment = a(h="right"); stlbl.border = b()
        tot = ws.cell(row=ROW, column=6,
            value=f"=SUMIF(A{st}:A{ROW-1},\"{current_loc}\",F{st}:F{ROW-1})")
        tot.number_format = JPY; tot.font = f(bold=True, sz=12)
        tot.fill = fill(LOC_HD_COLORS[current_loc]); tot.alignment = a(h="right"); tot.border = b()
        for col in [7,8,9]:
            ws.cell(row=ROW, column=col).fill = fill(LOC_HD_COLORS[current_loc])
            ws.cell(row=ROW, column=col).border = b()
        ROW += 2

    # ── 総合計サマリー ──────────────────────────────────
    ws.row_dimensions[ROW].height = 10
    ROW += 1
    ws.merge_cells(f"A{ROW}:I{ROW}")
    sep = ws["A" + str(ROW)]
    sep.fill = fill(NAVY); sep.border = b()
    ROW += 1

    ws.row_dimensions[ROW].height = 28
    ws.merge_cells(f"A{ROW}:E{ROW}")
    slbl = ws.cell(row=ROW, column=1, value="▶ 月額人件費 総合計（全拠点）")
    slbl.font = f(bold=True, sz=13, color=WHITE)
    slbl.fill = fill(NAVY); slbl.alignment = a(h="center"); slbl.border = b()

    all_f_cells = [r for r in range(6, ROW) if ws.cell(r, 1).value in LOCS]
    total_formula = f"=SUMIF(A6:A{ROW-2},\"<>管理\",F6:F{ROW-2})"
    stot = ws.cell(row=ROW, column=6, value=total_formula)
    stot.number_format = JPY; stot.font = f(bold=True, sz=14)
    stot.fill = fill("C8F7D4"); stot.alignment = a(h="right"); stot.border = b()
    for col in [7,8]:
        ws.cell(row=ROW, column=col).fill = fill(NAVY); ws.cell(row=ROW, column=col).border = b()
    ws.cell(row=ROW, column=9, value=f"=F{ROW}*12")
    ws.cell(row=ROW, column=9).number_format = JPY
    ws.cell(row=ROW, column=9).font = f(bold=True, sz=11, color=GRAY)
    ws.cell(row=ROW, column=9).fill = fill(LGRAY); ws.cell(row=ROW, column=9).border = b()
    ROW += 1

    # 注意書き
    ws.merge_cells(f"A{ROW}:I{ROW}")
    note2 = ws.cell(row=ROW, column=1,
        value="※ 上記は月額固定部分のみ。社会保険料（会社負担・約15%）・賞与・退職積立等は別途加算が必要です。")
    note2.font = f(sz=9, color="C0392B"); note2.fill = fill(REDLT)
    note2.alignment = a(h="left"); note2.border = b()
    ws.row_dimensions[ROW].height = 18


# ════════════════════════════════════════════════════════════════
#  BUILD
# ════════════════════════════════════════════════════════════════
def build():
    print("データ読込中...")
    sales = load_sales()
    print(f"  売上データ: {len(sales)}ヶ月")

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet()  # 案件別採算
    ws2 = wb.create_sheet()  # ルート別採算
    ws3 = wb.create_sheet()  # ▼売上原票
    ws4 = wb.create_sheet()  # ▼マッピング設定
    ws5 = wb.create_sheet()  # ▼従業員マスター

    sheet_anken(ws1, sales)
    sheet_route(ws2, sales)
    sheet_uriagehyo(ws3, sales)
    sheet_mapping(ws4)
    sheet_employee_master(ws5)

    ws1.sheet_properties.tabColor = NAVY
    ws2.sheet_properties.tabColor = "1A6B2E"
    ws3.sheet_properties.tabColor = "335577"
    ws4.sheet_properties.tabColor = "335577"
    ws5.sheet_properties.tabColor = "795548"

    out = "/home/user/sacure/sacure_採算管理_v4_ルート別案件別.xlsx"
    wb.save(out)
    print(f"保存: {out}")
    for ws_ in wb.worksheets:
        print(f"  {ws_.title}")

if __name__ == "__main__":
    build()
