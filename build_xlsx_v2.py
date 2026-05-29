"""
民間廃棄物収集受託案件 採算管理ブック v2
- 実際の拠点マスター（505拠点）を取り込み
- 3案件（東急/都営/一般）に基づいた採算管理
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import DataPoint
import datetime
from collections import defaultdict

SRC = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/83a04c7f-__________________.xlsx"

# ── Palette ────────────────────────────────────────────────────
C_NAVY   = "0D2B5E"
C_BLUE   = "1A5FC8"
C_GREEN  = "2EB872"
C_LGRAY  = "F2F6FC"
C_DGRAY  = "55657A"
C_WHITE  = "FFFFFF"
C_YELLOW = "FFF3CD"
C_RED_LT = "FFE0E0"
C_GRN_LT = "D4EDDA"
C_BLU_LT = "D0E8FF"
C_BORDER = "C5D5E8"

# Case colors
CASE_COLORS = {
    "東急グループ":  ("1565A8", "D0E8FF"),   # blue
    "都営地下鉄":    ("5B2D8E", "EDE7F6"),   # purple
    "一般事業者":    ("1A6B2E", "D4EDDA"),   # green
}

def fill(h): return PatternFill("solid", fgColor=h)
def font(bold=False, size=11, color=C_NAVY, name="Yu Gothic UI"):
    return Font(bold=bold, size=size, color=color, name=name)
def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def thin_border():
    s = Side(style="thin", color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def sc(ws, row, col, value=None, bold=False, size=11, fg=None,
       fc=C_NAVY, ha="left", va="center", wrap=False, fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font      = font(bold=bold, size=size, color=fc)
    c.alignment = align(h=ha, v=va, wrap=wrap)
    if fg: c.fill = fill(fg)
    if fmt: c.number_format = fmt
    c.border = thin_border()
    return c

def title_row(ws, text, cols, row=1, height=36, bg=C_NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font      = Font(bold=True, size=15, color=C_WHITE, name="Yu Gothic UI")
    c.fill      = fill(bg)
    c.alignment = align(h="center")
    ws.row_dimensions[row].height = height

def section_header(ws, row, texts_colors, h=22):
    """texts_colors = [(text, start_col, end_col, bg_hex), ...]"""
    ws.row_dimensions[row].height = h
    for text, sc_, ec_, bg in texts_colors:
        ws.merge_cells(start_row=row, start_column=sc_, end_row=row, end_column=ec_)
        c = ws.cell(row=row, column=sc_, value=text)
        c.font      = Font(bold=True, size=10, color=C_WHITE, name="Yu Gothic UI")
        c.fill      = fill(bg)
        c.alignment = align(h="center")
        c.border    = thin_border()

def col_headers(ws, row, headers, h=22):
    """headers = [(label, width, bg), ...]"""
    ws.row_dimensions[row].height = h
    for col, (label, width, bg) in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
        sc(ws, row, col, label, bold=True, size=9, fg=bg, fc=C_WHITE, ha="center")

# ── Load source data ────────────────────────────────────────────
def load_source():
    wb = openpyxl.load_workbook(SRC, data_only=True)

    # 拠点マスター
    ws_m = wb["①拠点マスター"]
    locations = []
    for row in ws_m.iter_rows(min_row=3, values_only=True):
        code, name, cat, addr, day, driver, wtype, kg = (list(row)+[None]*8)[:8]
        if code is None or code == "拠点コード":
            continue
        try: kg = float(kg) if kg not in (None,"") else 0
        except: kg = 0
        locations.append({
            "code": code, "name": name, "cat": cat or "",
            "addr": addr or "", "day": day or "",
            "driver": driver or "", "wtype": wtype or "", "kg_week": kg,
        })

    # 換算テーブル
    ws_cv = wb["②換算テーブル"]
    conv = []
    for row in ws_cv.iter_rows(min_row=3, values_only=True):
        if row[0] and not str(row[0]).startswith("【"):
            conv.append(row[:5])

    # ドライバー集計
    ws_dr = wb["⑥ドライバー別集計"]
    drivers = []
    for row in ws_dr.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] != "合計":
            drivers.append(row[:6])

    return locations, conv, drivers

# ════════════════════════════════════════════════════════════════
#  SHEET 1: 案件マスタ（3案件）
# ════════════════════════════════════════════════════════════════
def build_anken_master(ws, locations):
    ws.title = "①案件マスタ"
    ws.sheet_view.showGridLines = False

    title_row(ws, "民間廃棄物収集受託案件　案件マスタ", 12)

    # 案件サマリー from actual data
    from collections import defaultdict
    cats = defaultdict(lambda: {"count":0,"kg":0,"drivers":set(),"codes":[]})
    for loc in locations:
        cat = loc["cat"]
        cats[cat]["count"] += 1
        cats[cat]["kg"] += loc["kg_week"]
        for d in str(loc["driver"]).split("・"):
            if d.strip(): cats[cat]["drivers"].add(d.strip())
        cats[cat]["codes"].append(loc["code"])

    # Section header
    section_header(ws, 2, [
        ("案件情報", 1, 6, C_NAVY),
        ("規模・拠点", 7, 9, "1565A8"),
        ("月次売上目標（入力欄）", 10, 12, "1A6B2E"),
    ])

    hdrs = [
        ("案件ID",9), ("案件名",22), ("カテゴリ",14), ("担当拠点プレフィックス",18),
        ("担当ドライバー数",16), ("ステータス",12),
        ("拠点数",10), ("週間回収量\n(kg)",12), ("月間回収量\n(kg推算)",14),
        ("月次売上\n目標(円)",16), ("km単価\n(円/km)",12), ("備考",18),
    ]
    col_headers(ws, 3, [(h,w,C_BLUE) for h,w in hdrs])
    ws.row_dimensions[3].height = 30

    cases = [
        ("C001", "東急グループ案件",  "東急グループ", "TK-"),
        ("C002", "都営地下鉄案件",    "都営地下鉄",   "MT-"),
        ("C003", "一般事業者案件",    "一般事業者",   "GN-"),
    ]

    for r, (cid, cname, cat, prefix) in enumerate(cases, 4):
        ws.row_dimensions[r].height = 30
        d = cats[cat]
        bg_fg, bg_lt = CASE_COLORS.get(cat, (C_BLUE, C_BLU_LT))
        bg = bg_lt

        sc(ws, r, 1,  cid,   bold=True, size=11, fg=bg, fc=C_NAVY, ha="center")
        sc(ws, r, 2,  cname, bold=True, size=11, fg=bg, fc=C_NAVY)
        c = ws.cell(row=r, column=3, value=cat)
        c.font = Font(bold=True, size=11, color=C_WHITE, name="Yu Gothic UI")
        c.fill = fill(bg_fg); c.alignment = align(h="center"); c.border = thin_border()
        sc(ws, r, 4,  prefix,          size=11, fg=bg, ha="center")
        sc(ws, r, 5,  len(d["drivers"]),size=11, fg=bg, ha="center", fmt="#,##0")
        sc(ws, r, 6,  "稼働中",         size=11, fg=bg, ha="center")
        sc(ws, r, 7,  d["count"],       size=12, bold=True, fg=bg, ha="center", fmt="#,##0")
        sc(ws, r, 8,  d["kg"],          size=12, bold=True, fg=bg, ha="right",  fmt="#,##0")
        c9 = ws.cell(row=r, column=9)
        c9.value = f"=H{r}*4.33"; c9.number_format = "#,##0"; c9.font = font(bold=True, size=12)
        c9.fill = fill(bg); c9.alignment = align(h="right"); c9.border = thin_border()
        # Input cells (green tint)
        sc(ws, r, 10, 0, size=11, fg="E8F5E9", ha="right", fmt="#,##0")
        sc(ws, r, 11, 0, size=11, fg="E8F5E9", ha="right", fmt="#,##0")
        sc(ws, r, 12, "", size=10, fg=bg)

    # Total row
    tr = 7
    ws.row_dimensions[tr].height = 26
    sc(ws, tr, 1, "合　計", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center")
    for c2 in range(2, 6):
        sc(ws, tr, c2, "", bold=True, fg=C_NAVY, fc=C_WHITE)
    sc(ws, tr, 7, f"=SUM(G4:G6)", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center", fmt="#,##0")
    sc(ws, tr, 8, f"=SUM(H4:H6)", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="right",  fmt="#,##0")
    sc(ws, tr, 9, f"=SUM(I4:I6)", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="right",  fmt="#,##0")
    sc(ws, tr, 10,f"=SUM(J4:J6)",bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="right",  fmt="#,##0")
    for c2 in [11, 12]:
        sc(ws, tr, c2, "", bold=True, fg=C_NAVY, fc=C_WHITE)

    # Note
    ws.row_dimensions[9].height = 18
    ws.merge_cells("A9:L9")
    n = ws.cell(row=9, column=1, value="※ 月次売上目標・km単価はJ列・K列に入力してください。月次採算入力シートと連動します。")
    n.font = Font(size=9, italic=True, color=C_DGRAY, name="Yu Gothic UI")
    n.fill = fill(C_LGRAY); n.alignment = align(h="left"); n.border = thin_border()

# ════════════════════════════════════════════════════════════════
#  SHEET 2: 拠点マスター（505拠点をそのまま取込）
# ════════════════════════════════════════════════════════════════
def build_kyoten_master(ws, locations):
    ws.title = "②拠点マスター"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A4"

    title_row(ws, f"拠点マスター（{len(locations)}拠点）　※自動取込", 11)

    section_header(ws, 2, [
        ("拠点基本情報", 1, 8, C_NAVY),
        ("採算参考値（入力欄）", 9, 11, "1A6B2E"),
    ])

    hdrs = [
        ("拠点コード",11),("拠点名",28),("カテゴリ",14),("住所",30),
        ("回収曜日",14),("担当ドライバー",20),("廃棄物種別",22),("週間回収量\n(kg)",12),
        ("月間回収量\n(kg推算)",14),("収集単価\n(円/kg)",12),("月間収益\n目安(円)",14),
    ]
    col_headers(ws, 3, [(h,w,C_BLUE) for h,w in hdrs])
    ws.row_dimensions[3].height = 30

    cat_bg = {
        "東急グループ": CASE_COLORS["東急グループ"][1],
        "都営地下鉄":   CASE_COLORS["都営地下鉄"][1],
        "一般事業者":   CASE_COLORS["一般事業者"][1],
    }

    for r, loc in enumerate(locations, 4):
        ws.row_dimensions[r].height = 18
        bg = cat_bg.get(loc["cat"], C_WHITE)
        ha_c = "center"

        sc(ws, r, 1,  loc["code"],    size=10, fg=bg, ha="center")
        sc(ws, r, 2,  loc["name"],    size=10, fg=bg, ha="left", wrap=True)
        c3 = ws.cell(row=r, column=3, value=loc["cat"])
        c3.font = Font(size=10, color=C_WHITE, name="Yu Gothic UI")
        c3.fill = fill(CASE_COLORS.get(loc["cat"], (C_DGRAY,""))[0])
        c3.alignment = align(h="center", v="center"); c3.border = thin_border()
        sc(ws, r, 4,  loc["addr"],    size=9,  fg=bg, ha="left",  wrap=True)
        sc(ws, r, 5,  loc["day"],     size=9,  fg=bg, ha="center")
        sc(ws, r, 6,  loc["driver"],  size=9,  fg=bg, ha="left",  wrap=True)
        sc(ws, r, 7,  loc["wtype"],   size=9,  fg=bg, ha="left",  wrap=True)
        sc(ws, r, 8,  loc["kg_week"], size=10, bold=True, fg=bg, ha="right", fmt="#,##0")
        # 月間回収量 = 週間×4.33
        c9 = ws.cell(row=r, column=9)
        c9.value = f"=ROUND(H{r}*4.33,0)" if loc["kg_week"] > 0 else 0
        c9.number_format = "#,##0"; c9.font = font(size=10)
        c9.fill = fill(bg); c9.alignment = align(h="right"); c9.border = thin_border()
        # 入力欄（収集単価）
        sc(ws, r, 10, 0, size=10, fg="FAFFF5", ha="right", fmt="#,##0")
        # 月間収益
        c11 = ws.cell(row=r, column=11)
        c11.value = f"=IF(J{r}>0,I{r}*J{r},\"\")"
        c11.number_format = "#,##0"; c11.font = font(size=10, bold=True)
        c11.fill = fill("FAFFF5"); c11.alignment = align(h="right"); c11.border = thin_border()

    # カテゴリ別小計
    last = 3 + len(locations)
    sub_r = last + 2
    ws.row_dimensions[sub_r].height = 22
    sc(ws, sub_r, 1, "カテゴリ別集計", bold=True, size=11, fg=C_NAVY, fc=C_WHITE, ha="center")
    ws.merge_cells(f"A{sub_r}:C{sub_r}")
    for i, (cat, (fg_c, bg_c)) in enumerate(CASE_COLORS.items()):
        r2 = sub_r + i + 1
        ws.row_dimensions[r2].height = 22
        c = ws.cell(row=r2, column=1, value=cat)
        c.font = Font(bold=True, size=11, color=C_WHITE, name="Yu Gothic UI")
        c.fill = fill(fg_c); c.alignment = align(h="center"); c.border = thin_border()
        ws.merge_cells(f"A{r2}:C{r2}")
        # COUNTIF / SUMIF
        c_cnt = ws.cell(row=r2, column=7, value=f'=COUNTIF(C4:C{last},"{cat}")拠点')
        c_kg  = ws.cell(row=r2, column=8, value=f'=SUMIF(C4:C{last},"{cat}",H4:H{last})')
        for cc in [c_cnt, c_kg]:
            cc.font = font(size=11, bold=True); cc.fill = fill(bg_c)
            cc.alignment = align(h="right"); cc.border = thin_border()
        c_kg.number_format = "#,##0"

# ════════════════════════════════════════════════════════════════
#  SHEET 3: 月次採算入力
# ════════════════════════════════════════════════════════════════
def build_monthly_input(ws):
    ws.title = "③月次採算入力"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "E4"

    title_row(ws, "月次採算入力シート（案件別）", 20)

    section_header(ws, 2, [
        ("案件情報", 1, 4, C_NAVY),
        ("売　上", 5, 9, "1565A8"),
        ("原　価", 10, 17, "7B3F00"),
        ("採　算", 18, 21, "1A6B2E"),
    ])

    hdrs = [
        ("案件ID",9),("案件名",22),("カテゴリ",14),("集計年月",11),
        ("基本料金\n(円)",13),("スポット\n料金(円)",13),("処理委託\n収入(円)",12),("その他\n売上(円)",11),("売上\n合計(円)",14),
        ("ドライバー\n人件費(円)",14),("燃料費\n(円)",12),("車両\n維持費(円)",12),("中間\n処理費(円)",12),
        ("搬入・\n運搬費(円)",12),("電子M\n代行費(円)",12),("管理費\n(円)",11),("原価\n合計(円)",14),
        ("粗利益\n(円)",14),("粗利率",10),("採算\n判定",10),("備考",16),
    ]
    col_headers(ws, 3, [(h,w,C_BLUE) for h,w in hdrs])
    ws.row_dimensions[3].height = 36

    JPY = "#,##0"
    PCT = "0.0%"

    # 実績サンプルデータ（3案件×直近3ヶ月）
    samples = [
        # C001 東急グループ
        ("C001","東急グループ案件","東急グループ","2026-04",
         2850000,320000,180000,50000,
         680000,185000,95000,620000,85000,45000,120000),
        ("C001","東急グループ案件","東急グループ","2026-05",
         2850000,280000,175000,30000,
         675000,178000,92000,608000,82000,43000,118000),
        ("C001","東急グループ案件","東急グループ","2026-06",
         2850000,410000,190000,60000,
         690000,195000,98000,635000,90000,47000,125000),
        # C002 都営地下鉄
        ("C002","都営地下鉄案件","都営地下鉄","2026-04",
         1950000,80000,95000,20000,
         420000,125000,68000,380000,45000,30000,85000),
        ("C002","都営地下鉄案件","都営地下鉄","2026-05",
         1950000,65000,92000,15000,
         415000,120000,65000,372000,42000,28000,82000),
        # C003 一般事業者
        ("C003","一般事業者案件","一般事業者","2026-04",
         5200000,520000,410000,80000,
         1350000,485000,220000,1180000,165000,85000,210000),
        ("C003","一般事業者案件","一般事業者","2026-05",
         5200000,480000,398000,60000,
         1340000,472000,215000,1155000,160000,82000,205000),
        ("C003","一般事業者案件","一般事業者","2026-06",
         5200000,610000,425000,90000,
         1365000,502000,228000,1205000,172000,88000,218000),
    ]

    for r, (cid, cname, cat, ym, basic, spot, processing, other_s,
            driver, fuel, veh, proc_fee, haul, emani, mgmt) in enumerate(samples, 4):
        ws.row_dimensions[r].height = 22
        bg = CASE_COLORS.get(cat, (C_BLUE, C_BLU_LT))[1]

        sc(ws, r,  1, cid,    size=10, fg=bg, ha="center", bold=True)
        sc(ws, r,  2, cname,  size=10, fg=bg)
        c3 = ws.cell(row=r, column=3, value=cat)
        c3.font = Font(size=10, color=C_WHITE, name="Yu Gothic UI")
        c3.fill = fill(CASE_COLORS.get(cat,(C_BLUE,""))[0])
        c3.alignment = align(h="center"); c3.border = thin_border()
        sc(ws, r,  4, ym,     size=10, fg=bg, ha="center")

        # 売上列
        for col, val in zip(range(5,9), [basic, spot, processing, other_s]):
            sc(ws, r, col, val, size=10, fg=C_BLU_LT, ha="right", fmt=JPY)
        c9 = ws.cell(row=r, column=9)
        c9.value = f"=E{r}+F{r}+G{r}+H{r}"
        c9.number_format = JPY; c9.font = font(bold=True, size=11)
        c9.fill = fill(C_BLU_LT); c9.alignment = align(h="right"); c9.border = thin_border()

        # 原価列
        for col, val in zip(range(10,17), [driver, fuel, veh, proc_fee, haul, emani, mgmt]):
            sc(ws, r, col, val, size=10, fg=C_YELLOW, ha="right", fmt=JPY)
        c17 = ws.cell(row=r, column=17)
        c17.value = f"=J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r}"
        c17.number_format = JPY; c17.font = font(bold=True, size=11)
        c17.fill = fill(C_YELLOW); c17.alignment = align(h="right"); c17.border = thin_border()

        # 採算列
        c18 = ws.cell(row=r, column=18)
        c18.value = f"=I{r}-Q{r}"
        c18.number_format = JPY; c18.font = font(bold=True, size=11)
        c18.fill = fill(C_GRN_LT); c18.alignment = align(h="right"); c18.border = thin_border()

        c19 = ws.cell(row=r, column=19)
        c19.value = f"=IF(I{r}<>0,R{r}/I{r},0)"
        c19.number_format = PCT; c19.font = font(bold=True, size=11)
        c19.fill = fill(C_GRN_LT); c19.alignment = align(h="center"); c19.border = thin_border()

        c20 = ws.cell(row=r, column=20)
        c20.value = f'=IF(S{r}>=0.25,"◎ 優良",IF(S{r}>=0.15,"○ 良好",IF(S{r}>=0.05,"△ 要注意","× 赤字")))'
        c20.font = font(bold=True, size=10); c20.fill = fill(C_GRN_LT)
        c20.alignment = align(h="center"); c20.border = thin_border()

        sc(ws, r, 21, "", size=10, fg=bg)

    # 空白入力行 20行
    last_sample = 3 + len(samples)
    for i in range(20):
        r = last_sample + 1 + i
        ws.row_dimensions[r].height = 22
        bg = C_LGRAY if r % 2 == 0 else C_WHITE
        for col in range(1, 22):
            sc(ws, r, col, None, size=10, fg=bg,
               ha="right" if col >= 5 else ("center" if col in (1,3,4) else "left"),
               fmt=JPY if 5 <= col <= 17 else None)
        # formulas
        for col, expr in [
            (9,  f"=IF(E{r}+F{r}+G{r}+H{r}=0,\"\",E{r}+F{r}+G{r}+H{r})"),
            (17, f"=IF(J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r}=0,\"\",J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r})"),
            (18, f'=IF(OR(I{r}="",I{r}=0),"",I{r}-Q{r})'),
            (19, f'=IF(OR(I{r}="",I{r}=0),"",R{r}/I{r})'),
            (20, f'=IF(S{r}="","",IF(S{r}>=0.25,"◎ 優良",IF(S{r}>=0.15,"○ 良好",IF(S{r}>=0.05,"△ 要注意","× 赤字"))))'),
        ]:
            c2 = ws.cell(row=r, column=col)
            c2.value = expr
            c2.number_format = PCT if col == 19 else (JPY if col in (9,17,18) else "@")
            c2.font = font(bold=(col in (9,17,18,19,20)), size=10)
            c2.fill = fill(C_BLU_LT if col in (5,6,7,8,9) else
                           (C_YELLOW if col in (10,11,12,13,14,15,16,17) else
                            (C_GRN_LT if col in (18,19,20) else bg)))
            c2.alignment = align(h="center" if col == 20 else "right")
            c2.border = thin_border()

    # 条件付き書式（採算判定）
    last_r = last_sample + 20
    ws.conditional_formatting.add(f"T4:T{last_r}",
        FormulaRule(formula=[f'T4="◎ 優良"'],  fill=fill("C8F7D4"), font=Font(bold=True, color="145A32")))
    ws.conditional_formatting.add(f"T4:T{last_r}",
        FormulaRule(formula=[f'T4="○ 良好"'],  fill=fill(C_GRN_LT), font=Font(bold=True, color="1A6B2E")))
    ws.conditional_formatting.add(f"T4:T{last_r}",
        FormulaRule(formula=[f'T4="△ 要注意"'], fill=fill(C_YELLOW),  font=Font(bold=True, color="856404")))
    ws.conditional_formatting.add(f"T4:T{last_r}",
        FormulaRule(formula=[f'T4="× 赤字"'],  fill=fill(C_RED_LT),  font=Font(bold=True, color="8B0000")))
    ws.conditional_formatting.add(f"S4:S{last_r}",
        ColorScaleRule(start_type="num", start_value=0, start_color="FF6B6B",
                       mid_type="num", mid_value=0.15, mid_color="FFD93D",
                       end_type="num", end_value=0.3, end_color="6BCB77"))

# ════════════════════════════════════════════════════════════════
#  SHEET 4: 案件別月次集計
# ════════════════════════════════════════════════════════════════
def build_case_summary(ws):
    ws.title = "④案件別月次集計"
    ws.sheet_view.showGridLines = False

    title_row(ws, "案件別月次集計（2026年4月〜）", 14)

    cases = [
        ("C001", "東急グループ案件",  "東急グループ"),
        ("C002", "都営地下鉄案件",    "都営地下鉄"),
        ("C003", "一般事業者案件",    "一般事業者"),
    ]
    months = ["2026-04","2026-05","2026-06","2026-07","2026-08","2026-09"]

    # 集計データ（③シートのデータをベースに手動入力）
    data = {
        ("C001","2026-04"): (3400000,1630000),
        ("C001","2026-05"): (3335000,1598000),
        ("C001","2026-06"): (3510000,1680000),
        ("C002","2026-04"): (2145000,1153000),
        ("C002","2026-05"): (2122000,1124000),
        ("C003","2026-04"): (6210000,3695000),
        ("C003","2026-05"): (6138000,3629000),
        ("C003","2026-06"): (6325000,3776000),
    }

    row = 2
    for cid, cname, cat in cases:
        fg_c, bg_c = CASE_COLORS.get(cat, (C_BLUE, C_BLU_LT))
        ws.row_dimensions[row].height = 28
        ws.merge_cells(f"A{row}:N{row}")
        c = ws.cell(row=row, column=1, value=f"  {cid}  {cname}")
        c.font = Font(bold=True, size=13, color=C_WHITE, name="Yu Gothic UI")
        c.fill = fill(fg_c); c.alignment = align(h="left"); c.border = thin_border()

        row += 1
        ws.row_dimensions[row].height = 22
        hdrs2 = ["集計月","売上合計","原価合計","粗利益","粗利率","前月比 売上","前月比 粗利","判定"]
        widths = [12,16,16,16,10,14,14,12]
        for col, (h, w) in enumerate(zip(hdrs2, widths), 1):
            ws.column_dimensions[get_column_letter(col)].width = w
            sc(ws, row, col, h, bold=True, size=10, fg=C_NAVY, fc=C_WHITE, ha="center")

        hdr_row = row
        row += 1
        first_data_row = row
        for ym in months:
            ws.row_dimensions[row].height = 22
            sales, cost = data.get((cid, ym), (None, None))
            gross = (sales - cost) if sales and cost else None
            rate  = (gross / sales) if gross and sales else None

            sc(ws, row, 1, ym,    size=11, fg=bg_c, ha="center")
            sc(ws, row, 2, sales, size=11, fg=bg_c, ha="right", fmt="#,##0")
            sc(ws, row, 3, cost,  size=11, fg=bg_c, ha="right", fmt="#,##0")
            sc(ws, row, 4, gross, size=11, bold=True, fg=bg_c, ha="right", fmt="#,##0")
            sc(ws, row, 5, rate,  size=11, bold=True, fg=bg_c, ha="center", fmt="0.0%")
            if row == first_data_row:
                sc(ws, row, 6, "—", size=10, fg=bg_c, ha="center")
                sc(ws, row, 7, "—", size=10, fg=bg_c, ha="center")
            else:
                for col, ref in [(6, "B"), (7, "D")]:
                    c2 = ws.cell(row=row, column=col)
                    c2.value = f"=IF({ref}{row-1}<>0,{ref}{row}/{ref}{row-1}-1,\"\")"
                    c2.number_format = "+0.0%;-0.0%;0.0%"
                    c2.font = font(size=10); c2.fill = fill(bg_c)
                    c2.alignment = align(h="center"); c2.border = thin_border()
            verdict = ("◎ 優良" if (rate or 0) >= 0.25 else
                       "○ 良好" if (rate or 0) >= 0.15 else
                       "△ 要注意" if (rate or 0) >= 0.05 else "—" if rate is None else "× 赤字")
            vbg = {"◎ 優良":"C8F7D4","○ 良好":C_GRN_LT,"△ 要注意":C_YELLOW,"× 赤字":C_RED_LT,"—":bg_c}
            sc(ws, row, 8, verdict, size=10, bold=True, fg=vbg[verdict],
               fc=("145A32" if verdict in ("◎ 優良","○ 良好") else "856404" if verdict == "△ 要注意" else "8B0000" if verdict == "× 赤字" else C_NAVY),
               ha="center")
            row += 1

        # Sub-total
        sub_r = row
        ws.row_dimensions[sub_r].height = 22
        for col, expr, fmt in [
            (2, f"=SUM(B{first_data_row}:B{sub_r-1})", "#,##0"),
            (3, f"=SUM(C{first_data_row}:C{sub_r-1})", "#,##0"),
            (4, f"=SUM(D{first_data_row}:D{sub_r-1})", "#,##0"),
            (5, f"=IF(B{sub_r}<>0,D{sub_r}/B{sub_r},0)", "0.0%"),
        ]:
            c2 = ws.cell(row=sub_r, column=col)
            c2.value = expr; c2.number_format = fmt
            c2.font = Font(bold=True, size=11, color=C_WHITE, name="Yu Gothic UI")
            c2.fill = fill(fg_c); c2.alignment = align(h="right"); c2.border = thin_border()
        sc(ws, sub_r, 1, "小　計", bold=True, size=11, fg=fg_c, fc=C_WHITE, ha="center")
        for col in [6,7,8]: sc(ws, sub_r, col, "", fg=fg_c, fc=C_WHITE, bold=True)
        row += 2

# ════════════════════════════════════════════════════════════════
#  SHEET 5: ドライバー別採算
# ════════════════════════════════════════════════════════════════
def build_driver(ws, drivers):
    ws.title = "⑤ドライバー別採算"
    ws.sheet_view.showGridLines = False

    title_row(ws, "ドライバー別　回収実績・コスト配賦", 11)

    section_header(ws, 2, [
        ("ドライバー情報", 1, 5, C_NAVY),
        ("回収実績", 6, 8, "1565A8"),
        ("コスト（月次・入力欄）", 9, 11, "7B3F00"),
        ("効率指標", 12, 14, "1A6B2E"),
    ])

    hdrs = [
        ("ドライバー名",14),("担当拠点数",12),("主担当案件",18),("雇用区分",12),("月稼働日数",12),
        ("週間回収量\n(kg)",12),("月間回収量\n(kg推算)",14),("担当拠点あたり\n(kg)",14),
        ("月次人件費\n(円)",14),("燃料費\n(円)",12),("その他\n費用(円)",12),
        ("1kg当り\nコスト(円)",12),("売上\n貢献額(円)",13),("採算\n貢献率",10),
    ]
    col_headers(ws, 3, [(h,w,C_BLUE) for h,w in hdrs])
    ws.row_dimensions[3].height = 36

    # ドライバー → 主担当案件マッピング
    driver_case = {
        "千葉": "C001/C003","中條": "C001/C003","長谷川": "C003","長野": "C001/C003",
        "宮原": "C002","岡田": "C002/C003","和田": "C001","荻野": "C001",
        "大迫": "C002","藤内": "C001","渡辺": "C002","柳沼": "C002",
        "玉山": "C002","白": "C003","鈴木": "C003","山口": "C003",
    }

    for r, dr_row in enumerate(drivers, 4):
        ws.row_dimensions[r].height = 22
        name = dr_row[0] if dr_row else ""
        try: spots = int(dr_row[1]) if dr_row[1] else 0
        except: spots = 0
        try: kg_week = float(dr_row[2]) if dr_row[2] else 0
        except: kg_week = 0
        try: kg_month = float(dr_row[3]) if dr_row[3] else 0
        except: kg_month = 0
        try: kg_per = float(dr_row[4]) if dr_row[4] else 0
        except: kg_per = 0

        bg = C_LGRAY if r % 2 == 0 else C_WHITE
        sc(ws, r, 1, name,  size=11, bold=True, fg=bg)
        sc(ws, r, 2, spots, size=11, fg=bg, ha="center", fmt="#,##0")
        sc(ws, r, 3, driver_case.get(str(name),"—"), size=10, fg=bg, ha="center")
        sc(ws, r, 4, "正社員" if spots > 100 else "パート", size=10, fg=bg, ha="center")
        sc(ws, r, 5, 22,    size=10, fg=bg, ha="center", fmt="#,##0")
        sc(ws, r, 6, kg_week,  size=11, fg=bg, ha="right", fmt="#,##0")
        sc(ws, r, 7, kg_month, size=11, fg=bg, ha="right", fmt="#,##0")
        sc(ws, r, 8, kg_per,   size=11, fg=bg, ha="right", fmt="#,##0.0")
        # コスト入力欄
        sc(ws, r, 9,  0, size=10, fg=C_YELLOW, ha="right", fmt="#,##0")
        sc(ws, r, 10, 0, size=10, fg=C_YELLOW, ha="right", fmt="#,##0")
        sc(ws, r, 11, 0, size=10, fg=C_YELLOW, ha="right", fmt="#,##0")
        # 効率指標
        c12 = ws.cell(row=r, column=12)
        c12.value = f"=IF(G{r}>0,(I{r}+J{r}+K{r})/G{r},0)"
        c12.number_format = "#,##0.0"; c12.font = font(bold=True, size=10)
        c12.fill = fill(C_GRN_LT); c12.alignment = align(h="right"); c12.border = thin_border()
        sc(ws, r, 13, 0, size=10, fg=C_GRN_LT, ha="right", fmt="#,##0")
        sc(ws, r, 14, 0, size=10, fg=C_GRN_LT, ha="right", fmt="0.0%")

    # 合計
    last = 3 + len(drivers)
    tr = last + 1
    ws.row_dimensions[tr].height = 24
    sc(ws, tr, 1, "合　計", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center")
    for col, col_l in [(2,"B"),(6,"F"),(7,"G"),(9,"I"),(10,"J"),(11,"K")]:
        tc = ws.cell(row=tr, column=col)
        tc.value = f"=SUM({col_l}4:{col_l}{last})"
        tc.number_format = "#,##0"; tc.font = Font(bold=True, size=12, color=C_WHITE, name="Yu Gothic UI")
        tc.fill = fill(C_NAVY); tc.alignment = align(h="right"); tc.border = thin_border()
    for col in [3,4,5,8,12,13,14]:
        sc(ws, tr, col, "", bold=True, fg=C_NAVY, fc=C_WHITE)

# ════════════════════════════════════════════════════════════════
#  SHEET 6: 採算ダッシュボード
# ════════════════════════════════════════════════════════════════
def build_dashboard(ws):
    ws.title = "⑥採算ダッシュボード"
    ws.sheet_view.showGridLines = False

    title_row(ws, "採算ダッシュボード（直近月：2026年6月）", 11)

    # KPI ボックス 3案件
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:E2")
    c = ws.cell(row=2, column=1, value="■ 案件別採算サマリー（2026年6月）")
    c.font = Font(bold=True, size=13, color=C_NAVY, name="Yu Gothic UI")
    c.alignment = align(h="left"); c.border = thin_border()
    for col in range(2,6): ws.cell(row=2, column=col).border = thin_border()

    kpi = [
        ("C001","東急グループ","東急グループ", 3510000, 1680000),
        ("C002","都営地下鉄",  "都営地下鉄",   0,       0),
        ("C003","一般事業者",  "一般事業者",   6325000, 3776000),
    ]

    col_widths = [10,22,16,16,16,14,12]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    hdrs2 = [("案件ID",C_NAVY),("案件名",C_NAVY),("売上",C_NAVY),("原価",C_NAVY),("粗利益",C_NAVY),("粗利率",C_NAVY),("判定",C_NAVY)]
    for col, (h, bg) in enumerate(hdrs2, 1):
        sc(ws, 3, col, h, bold=True, size=10, fg=C_NAVY, fc=C_WHITE, ha="center")
    ws.row_dimensions[3].height = 22

    for r, (cid, cname, cat, sales, cost) in enumerate(kpi, 4):
        fg_c, bg_c = CASE_COLORS.get(cat, (C_BLUE, C_BLU_LT))
        ws.row_dimensions[r].height = 28
        gross = sales - cost if sales else None
        rate  = gross / sales if gross and sales else None
        verdict = ("◎ 優良" if (rate or 0) >= 0.25 else
                   "○ 良好" if (rate or 0) >= 0.15 else
                   "— 未入力" if not sales else "△ 要注意")
        vbg = {"◎ 優良":"C8F7D4","○ 良好":C_GRN_LT,"△ 要注意":C_YELLOW,"— 未入力":C_LGRAY}
        sc(ws, r, 1, cid,    size=11, bold=True, fg=bg_c, ha="center")
        sc(ws, r, 2, cname,  size=12, bold=True, fg=bg_c)
        sc(ws, r, 3, sales,  size=12, fg=bg_c, ha="right",  fmt="#,##0")
        sc(ws, r, 4, cost,   size=12, fg=bg_c, ha="right",  fmt="#,##0")
        sc(ws, r, 5, gross,  size=12, bold=True, fg=bg_c, ha="right",  fmt="#,##0")
        sc(ws, r, 6, rate,   size=12, bold=True, fg=bg_c, ha="center", fmt="0.0%")
        sc(ws, r, 7, verdict, size=11, bold=True, fg=vbg.get(verdict, C_LGRAY),
           fc="145A32" if "優" in verdict else "1A6B2E" if "良" in verdict else C_NAVY, ha="center")

    # 合計行
    tr = 7
    ws.row_dimensions[tr].height = 26
    sc(ws, tr, 1, "合　計", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center")
    sc(ws, tr, 2, "",      bold=True, fg=C_NAVY, fc=C_WHITE)
    for col, l in [(3,"C"),(4,"D"),(5,"E")]:
        c2 = ws.cell(row=tr, column=col)
        c2.value = f"=SUM({l}4:{l}6)"; c2.number_format = "#,##0"
        c2.font = Font(bold=True, size=13, color=C_WHITE, name="Yu Gothic UI")
        c2.fill = fill(C_NAVY); c2.alignment = align(h="right"); c2.border = thin_border()
    c6 = ws.cell(row=tr, column=6)
    c6.value = "=IF(C7<>0,E7/C7,0)"; c6.number_format = "0.0%"
    c6.font = Font(bold=True, size=13, color=C_WHITE, name="Yu Gothic UI")
    c6.fill = fill(C_NAVY); c6.alignment = align(h="center"); c6.border = thin_border()
    sc(ws, tr, 7, "", bold=True, fg=C_NAVY, fc=C_WHITE)

    # 拠点規模サマリー
    ws.row_dimensions[9].height = 22
    ws.merge_cells("A9:G9")
    c9 = ws.cell(row=9, column=1, value="■ 拠点規模・週間回収量サマリー")
    c9.font = Font(bold=True, size=13, color=C_NAVY, name="Yu Gothic UI")
    c9.alignment = align(h="left"); c9.border = thin_border()

    scale_hdrs = ["カテゴリ","拠点数","週間回収量(kg)","月間回収量(kg推算)","全体シェア"]
    for col, h in enumerate(scale_hdrs, 1):
        sc(ws, 10, col, h, bold=True, size=10, fg=C_BLUE, fc=C_WHITE, ha="center")
    ws.row_dimensions[10].height = 22

    scale_data = [
        ("東急グループ", 93, 6312),
        ("都営地下鉄",  110, 8930),
        ("一般事業者",  302,37561),
    ]
    total_kg = sum(d[2] for d in scale_data)
    for r, (cat, cnt, kg) in enumerate(scale_data, 11):
        fg_c, bg_c = CASE_COLORS.get(cat, (C_BLUE, C_BLU_LT))
        ws.row_dimensions[r].height = 24
        c = ws.cell(row=r, column=1, value=cat)
        c.font = Font(bold=True, size=11, color=C_WHITE, name="Yu Gothic UI")
        c.fill = fill(fg_c); c.alignment = align(h="center"); c.border = thin_border()
        sc(ws, r, 2, cnt,        size=12, bold=True, fg=bg_c, ha="center", fmt="#,##0")
        sc(ws, r, 3, kg,         size=12, bold=True, fg=bg_c, ha="right",  fmt="#,##0")
        sc(ws, r, 4, round(kg*4.33), size=12, fg=bg_c, ha="right", fmt="#,##0")
        sc(ws, r, 5, kg/total_kg, size=12, fg=bg_c, ha="center", fmt="0.0%")
    # 合計
    sc(ws, 14, 1, "合　計", bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center")
    sc(ws, 14, 2, sum(d[1] for d in scale_data), bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center", fmt="#,##0")
    sc(ws, 14, 3, total_kg, bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="right", fmt="#,##0")
    sc(ws, 14, 4, round(total_kg*4.33), bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="right", fmt="#,##0")
    sc(ws, 14, 5, 1.0, bold=True, size=12, fg=C_NAVY, fc=C_WHITE, ha="center", fmt="0.0%")
    ws.row_dimensions[14].height = 26

    # Bar chart: 売上・粗利
    chart = BarChart()
    chart.type = "col"; chart.title = "案件別 売上・粗利（2026年6月）"
    chart.y_axis.title = "金額（円）"; chart.width = 20; chart.height = 12
    chart.grouping = "clustered"
    cats_ref = Reference(ws, min_col=2, min_row=4, max_row=6)
    sales_ref = Reference(ws, min_col=3, min_row=3, max_row=6)
    gross_ref = Reference(ws, min_col=5, min_row=3, max_row=6)
    chart.add_data(sales_ref, titles_from_data=True)
    chart.add_data(gross_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = C_BLUE
    chart.series[1].graphicalProperties.solidFill = C_GREEN
    ws.add_chart(chart, "A16")

# ════════════════════════════════════════════════════════════════
#  SHEET 7: 換算テーブル（元データ引継）
# ════════════════════════════════════════════════════════════════
def build_conv_table(ws, conv):
    ws.title = "⑦換算テーブル"
    ws.sheet_view.showGridLines = False

    title_row(ws, "品目×袋サイズ 重量換算テーブル（CO₂排出係数含む）", 9)

    hdrs = [("品目",16),("容器・サイズ",20),("1単位あたり重量(kg)",20),("備考",32),("CO₂排出係数\n(t-CO₂/t)",18)]
    col_headers(ws, 2, [(h,w,C_BLUE) for h,w in hdrs])
    ws.row_dimensions[2].height = 28

    for r, row_data in enumerate(conv, 3):
        ws.row_dimensions[r].height = 20
        bg = C_LGRAY if r % 2 == 0 else C_WHITE
        for col, val in enumerate(row_data, 1):
            fmt = "#,##0.0" if col == 3 else ("0.00000" if col == 5 else None)
            sc(ws, r, col, val, size=10, fg=bg, ha="right" if col in (3,5) else "left", fmt=fmt)

# ════════════════════════════════════════════════════════════════
#  BUILD
# ════════════════════════════════════════════════════════════════
def build():
    locations, conv, drivers = load_source()

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    ws1 = wb.create_sheet()
    ws2 = wb.create_sheet()
    ws3 = wb.create_sheet()
    ws4 = wb.create_sheet()
    ws5 = wb.create_sheet()
    ws6 = wb.create_sheet()
    ws7 = wb.create_sheet()

    build_anken_master(ws1, locations)
    build_kyoten_master(ws2, locations)
    build_monthly_input(ws3)
    build_case_summary(ws4)
    build_driver(ws5, drivers)
    build_dashboard(ws6)
    build_conv_table(ws7, conv)

    # Tab colors
    ws1.sheet_properties.tabColor = C_NAVY
    ws2.sheet_properties.tabColor = "1565A8"
    ws3.sheet_properties.tabColor = "7B3F00"
    ws4.sheet_properties.tabColor = "1A6B2E"
    ws5.sheet_properties.tabColor = "5B2D8E"
    ws6.sheet_properties.tabColor = "E07B00"
    ws7.sheet_properties.tabColor = C_DGRAY

    out = "/home/user/sacure/sacure_採算管理_民間廃棄物収集_v2.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print(f"  拠点数: {len(locations)}")
    print(f"  ドライバー数: {len(drivers)}")

if __name__ == "__main__":
    build()
