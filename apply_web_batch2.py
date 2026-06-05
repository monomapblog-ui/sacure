"""
WEB調査バッチ2: 確定エントリをD列に適用
赤文字+緑セル = WEB確認済み
赤文字+黄セル = 表記揺れ一致
"""
import openpyxl
from openpyxl.styles import PatternFill, Font

SAVED = '/home/user/sacure/サキュレ_拠点マスターv5_D列更新.xlsx'

green_fill  = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
red_font    = Font(color='FF0000')

# (C列キーワード, D列入力値, フィル種別)
ENTRIES = [
    ('鮮魚　北浜　学芸大学店', '共立食品（株）',                    'green'),
    ('鮮魚 北浜 学芸大学店',            '共立食品（株）',                    'green'),
    ('あーる工房',                       '（株）あ－る工房',                 'green'),
    ('オハナ　明大前',               '東急ウェルネス（株）',              'green'),
    ('オハナ 明大前',                    '東急ウェルネス（株）',              'green'),
    ('あおい皮膚科　駅ナカ上野毛',   '医）創青会',                       'green'),
    ('あおい皮膚科 駅ナカ上野毛',        '医）創青会',                       'green'),
    ('フラマンドール　田園調布店',   '（株）サンジェルマン',              'green'),
    ('フラマンドール 田園調布店',        '（株）サンジェルマン',              'green'),
    ('桜小路　共用部',               '東急（株）',                       'green'),
    ('桜小路 共用部',                    '東急（株）',                       'green'),
    ('ドトールコーヒー雪谷大塚店',       '（株）東急グルメフロント',          'green'),
    ('住まいと暮らしのコンシェルジュ',   '東急（株）',                       'green'),
    ('ケンタッキーフライドチキン　二子玉川店', '（株）東急グルメフロント', 'green'),
    ('ケンタッキーフライドチキン 二子玉川店',      '（株）東急グルメフロント', 'green'),
    ('ファットバー',                     'ファットバ－(グーポンズフードサービス', 'yellow'),
]

wb = openpyxl.load_workbook(SAVED)
ws = wb['Sheet1']

updated = []

for row in ws.iter_rows(min_row=2):
    c_cell = row[2]
    d_cell = row[3]

    if d_cell.value is not None:
        continue

    c_val = str(c_cell.value or '').strip()

    for keyword, billing_name, fill_type in ENTRIES:
        if c_val == keyword:
            d_cell.value = billing_name
            d_cell.font  = red_font
            d_cell.fill  = green_fill if fill_type == 'green' else yellow_fill
            updated.append((c_cell.row, c_val, billing_name, fill_type))
            break

wb.save(SAVED)
print(f"更新件数: {len(updated)}")
for r, c, d, t in updated:
    print(f"  行{r} [{t}]: {c} → {d}")
