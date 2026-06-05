"""
文脈推定バッチ: 東急（株）14件 + （有）花壇1件 → 赤文字・緑セル
"""
import openpyxl
from openpyxl.styles import PatternFill, Font

SAVED = '/home/user/sacure/サキュレ_拠点マスターv5_D列更新.xlsx'

green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
red_font   = Font(color='FF0000')

# 行番号 → 入力値（空欄のみ更新）
ROW_MAP = {
    185: '東急（株）',
    186: '東急（株）',
    206: '東急（株）',
    281: '東急（株）',
    295: '東急（株）',
    350: '東急（株）',
    351: '東急（株）',
    352: '東急（株）',
    368: '東急（株）',
    395: '東急（株）',
    417: '東急（株）',
    462: '東急（株）',
    483: '東急（株）',
    484: '東急（株）',
    386: '（有）花壇',
}

wb = openpyxl.load_workbook(SAVED)
ws = wb['Sheet1']

updated = []
for row in ws.iter_rows(min_row=2):
    r = row[0].row
    if r not in ROW_MAP:
        continue
    d_cell = row[3]
    if d_cell.value is not None:
        print(f"  スキップ 行{r}: すでに入力済み ({d_cell.value})")
        continue
    d_cell.value = ROW_MAP[r]
    d_cell.font  = red_font
    d_cell.fill  = green_fill
    updated.append((r, str(row[2].value), ROW_MAP[r]))

wb.save(SAVED)
print(f"更新件数: {len(updated)}")
for r, c, d in updated:
    print(f"  行{r}: {c} → {d}")
