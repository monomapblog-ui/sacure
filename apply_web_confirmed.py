"""
WEB確認済みエントリを D列に適用（赤文字 + 緑セル）
必ず保存済みファイルから読み込み、空欄のみ更新
"""
import openpyxl
from openpyxl.styles import PatternFill, Font

SAVED = '/home/user/sacure/サキュレ_拠点マスターv5_D列更新.xlsx'

green_fill = PatternFill(start_color='92D050', end_color='92D050', fill_type='solid')
red_font   = Font(color='FF0000')

# C列（収集先名）に含まれるキーワード → D列に入れる板橋B名称
WEB_MAP = [
    ('DEAN',            '（株）ウェルカム'),
    ('DEAN&DELUCA',     '（株）ウェルカム'),
    ('roobby',          '東急（株）'),
    ('緑が丘エキナカ保育園', '東急（株）'),
    ('ゴントラン',      '名鉄協商（株）'),
    ('TWG',             '（株）東急グルメフロント'),
    ('NATURAL KITCHEN', '（有）アミュ－ズフル'),
    ('DUMBO',           '（株）インセプション'),
    ('ルナアース',      '（株）エンドレス'),
    ('学大マーチ',      'アンシェフ'),
    ('E-style',         '（株）栄光'),
    ('大崎電気工業所',  '（株）大崎電機工業所'),
]

wb = openpyxl.load_workbook(SAVED)
ws = wb['Sheet1']

updated = []

for row in ws.iter_rows(min_row=2):
    c_cell = row[2]  # C列
    d_cell = row[3]  # D列

    if d_cell.value is not None:
        continue  # 既入力は触らない

    c_val = str(c_cell.value or '').strip()

    for keyword, billing_name in WEB_MAP:
        if keyword in c_val:
            d_cell.value = billing_name
            d_cell.font  = red_font
            d_cell.fill  = green_fill
            updated.append((c_cell.row, c_val, billing_name))
            break

wb.save(SAVED)
print(f"更新件数: {len(updated)}")
for row_no, c_name, d_name in updated:
    print(f"  行{row_no}: {c_name} → {d_name}")
