"""
民間廃棄物収集受託案件 採算管理ブック v3
- 売上原票シートを追加し VLOOKUP で採算入力と連携
"""
import openpyxl, xlrd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, FormulaRule
from openpyxl.chart import BarChart, Reference
from collections import defaultdict
import sys, os

SRC_MASTER = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/83a04c7f-__________________.xlsx"
SRC_2026   = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/912c64bc-___2026.xlsx"
SRC_R76    = "/root/.claude/uploads/ffa395c4-1170-42eb-9d67-c76992faca3c/8a0a54a8-__R7.6.xls"

# ── Palette ────────────────────────────────────────────────────
C_NAVY  = "0D2B5E";  C_BLUE  = "1A5FC8";  C_GREEN = "2EB872"
C_LGRAY = "F2F6FC";  C_DGRAY = "55657A";  C_WHITE = "FFFFFF"
C_YELLOW= "FFF3CD";  C_RED   = "FFE0E0";  C_GRN_LT= "D4EDDA"
C_BLU_LT= "D0E8FF";  C_BORDER= "C5D5E8"
CASE_BG = {"C001":"D0E8FF","C002":"EDE7F6","C003":"D4EDDA"}
CASE_FG = {"C001":"1565A8","C002":"5B2D8E","C003":"1A6B2E"}

def fill(h): return PatternFill("solid", fgColor=h)
def fnt(bold=False, sz=11, color=C_NAVY, name="Yu Gothic UI"):
    return Font(bold=bold, size=sz, color=color, name=name)
def aln(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def bdr():
    s = Side(style="thin", color=C_BORDER)
    return Border(top=s, bottom=s, left=s, right=s)

def sc(ws, r, c, val=None, bold=False, sz=11, bg=None, fc=C_NAVY,
       ha="left", va="center", wrap=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=val)
    cell.font      = fnt(bold=bold, sz=sz, color=fc)
    cell.alignment = aln(h=ha, v=va, wrap=wrap)
    cell.border    = bdr()
    if bg:  cell.fill = fill(bg)
    if fmt: cell.number_format = fmt
    return cell

def title_row(ws, txt, ncols, row=1, h=36, bg=C_NAVY):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=txt)
    c.font=fnt(bold=True,sz=15,color=C_WHITE); c.fill=fill(bg)
    c.alignment=aln(h="center"); ws.row_dimensions[row].height=h

def sec_hdr(ws, row, items, h=22):
    """items = [(text, col_from, col_to, bg), ...]"""
    ws.row_dimensions[row].height = h
    for txt, c1, c2, bg in items:
        ws.merge_cells(start_row=row,start_column=c1,end_row=row,end_column=c2)
        c = ws.cell(row=row,column=c1,value=txt)
        c.font=fnt(bold=True,sz=10,color=C_WHITE); c.fill=fill(bg)
        c.alignment=aln(h="center"); c.border=bdr()

def col_hdrs(ws, row, hdrs, h=22):
    """hdrs = [(label, width, bg), ...]"""
    ws.row_dimensions[row].height = h
    for i,(lbl,w,bg) in enumerate(hdrs,1):
        ws.column_dimensions[get_column_letter(i)].width = w
        sc(ws,row,i,lbl,bold=True,sz=9,bg=bg,fc=C_WHITE,ha="center")

# ════════════════════════════════════════════════════════════════
#  DATA EXTRACTION
# ════════════════════════════════════════════════════════════════
KAMOKU_CODES = [
    ("612-1","運送（本社）"),("612-2","入札（公共）"),("613-1","車輌"),
    ("613-2","KC"),("614-1","コンテナ"),("614-2","ENDO"),("614-3","現金"),
    ("615-1","板橋"),("615-2","新横浜"),("615-3","現金2"),("615-4","有価物"),
]
CODES = [k for k,_ in KAMOKU_CODES]

def reiwa_to_ym(s):
    s = str(s).strip()
    if not s.startswith("R"): return None
    try:
        era, m = s[1:].split(".")
        y = int(era) + 2018
        return f"{y}-{int(m):02d}"
    except: return None

def extract_sales(sales_data, ws_iter, is_openpyxl=True):
    if is_openpyxl:
        rows = [list(r) for r in ws_iter.iter_rows(min_row=3, values_only=True)]
        header = [str(v).strip() if v else "" for v in rows[0]]
        data_rows = rows[1:]
    else:  # xlrd
        ws = ws_iter
        header = [str(ws.cell_value(2,c)).strip() for c in range(ws.ncols)]
        data_rows = [[ws.cell_value(r,c) for c in range(ws.ncols)]
                     for r in range(3, ws.nrows)]

    for row in data_rows:
        if not row or not row[0] or not str(row[0]).startswith("R"): continue
        ym = reiwa_to_ym(str(row[0]))
        if not ym: continue
        d = {}
        for code in CODES:
            for ci, h in enumerate(header):
                if h == code and ci < len(row):
                    v = row[ci]
                    try: d[code] = float(v) if v and v != "" else 0.0
                    except: d[code] = 0.0
                    break
        # 合計(税抜) - 売上列
        for ci, h in enumerate(header):
            if "売" in h and "上" in h and ci < len(row):
                v = row[ci]
                try: d["_total"] = float(v) if v and v != "" else 0.0
                except: d["_total"] = 0.0
                break
        if any(v > 0 for v in d.values()):
            sales_data[ym] = d

def load_all_sales():
    sales = {}
    # R7.6.xls
    wb_old = xlrd.open_workbook(SRC_R76)
    extract_sales(sales, wb_old.sheet_by_name("売上"), is_openpyxl=False)
    # 2026.xlsx
    wb_new = openpyxl.load_workbook(SRC_2026, data_only=True)
    extract_sales(sales, wb_new["売上"], is_openpyxl=True)
    return {k: v for k, v in sorted(sales.items()) if v.get("_total", 0) > 0}

def load_locations():
    wb = openpyxl.load_workbook(SRC_MASTER, data_only=True)
    ws = wb["①拠点マスター"]
    locs = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        code, name, cat, addr, day, driver, wtype, kg = (list(row)+[None]*8)[:8]
        if not code or code == "拠点コード": continue
        try: kg = float(kg) if kg not in (None,"") else 0
        except: kg = 0
        locs.append(dict(code=code,name=name,cat=cat or "",
                         addr=addr or "",day=day or "",
                         driver=driver or "",wtype=wtype or "",kg_week=kg))
    return locs

def load_drivers():
    wb = openpyxl.load_workbook(SRC_MASTER, data_only=True)
    ws = wb["⑥ドライバー別集計"]
    rows = []
    for row in ws.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] not in ("合計","ドライバー"):
            rows.append(row[:6])
    return rows

# ════════════════════════════════════════════════════════════════
#  SHEET A: 売上原票
# ════════════════════════════════════════════════════════════════
def build_uriagehyo(ws, sales_data):
    ws.title = "⑤売上原票"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "C3"

    title_row(ws, "売上原票　月次実績（科目別・税抜）　※VLOOKUPキー列=A列(YYYY-MM)", 16)

    # 説明行
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:P2")
    c2 = ws.cell(row=2, column=1,
        value="【使い方】 ③月次採算入力の売上欄は =VLOOKUP(集計月, ⑤売上原票!$A:$P, 列番号, 0) で自動参照しています。毎月データを追記してください。")
    c2.font=fnt(sz=9,color=C_DGRAY); c2.fill=fill(C_LGRAY)
    c2.alignment=aln(h="left"); c2.border=bdr()

    # ヘッダー行
    hdrs = [
        ("年月\n(VLOOKUPKEY)",13),("R表記",9),
    ] + [(f"{c}\n{n}",13) for c,n in KAMOKU_CODES] + [
        ("合計\n(税抜)",15),("消費税\n(△)",12),("合計\n(税込)",15),
    ]
    # Section header
    sec_hdr(ws, 3, [
        ("VLOOKUP\nキー", 1, 2, C_NAVY),
        ("科　目　別　売　上　(円・税抜)", 3, 13, "1565A8"),
        ("合　計", 14, 16, "1A6B2E"),
    ], h=32)
    col_hdrs(ws, 4, [(lbl,w,C_BLUE) for lbl,w in hdrs], h=36)

    JPY = "#,##0"
    months = sorted(sales_data.keys())
    for r, ym in enumerate(months, 5):
        ws.row_dimensions[r].height = 20
        d = sales_data[ym]
        era_num = int(ym[:4]) - 2018
        era_str = f"R{era_num}.{int(ym[5:7])}"
        total_ex = d.get("_total", 0)
        total_inc = total_ex  # tax is negative in source, total(税込) = 合計(税抜) - 消費税負値
        # 消費税 ≈ total*0.1
        tax = -round(total_ex * 10/11)

        bg = C_LGRAY if r % 2 == 0 else C_WHITE
        sc(ws,r,1,ym,  bold=True,sz=11,bg=bg,ha="center")
        sc(ws,r,2,era_str,sz=10,bg=bg,ha="center")
        for ci,(code,_) in enumerate(KAMOKU_CODES,3):
            v = d.get(code,0)
            sc(ws,r,ci,v if v else None,sz=10,bg=bg,ha="right",fmt=JPY)
        sc(ws,r,14,total_ex if total_ex else None,bold=True,sz=11,bg=bg,ha="right",fmt=JPY)
        sc(ws,r,15,tax if total_ex else None,sz=10,bg=bg,ha="right",fmt=JPY)
        # 税込
        c16 = ws.cell(row=r,column=16)
        c16.value = f"=IF(N{r}<>0,N{r}+O{r},\"\")"
        c16.number_format=JPY; c16.font=fnt(bold=True,sz=11)
        c16.fill=fill(bg); c16.alignment=aln(h="right"); c16.border=bdr()

    # 将来月用 空白行（12ヶ月分）
    last = 4 + len(months)
    for i in range(12):
        r = last + 1 + i
        ws.row_dimensions[r].height = 20
        bg = C_LGRAY if r%2==0 else C_WHITE
        sc(ws,r,1,"",sz=11,bg=bg,ha="center")  # ← ここに YYYY-MM を入力
        sc(ws,r,2,"",sz=10,bg=bg,ha="center")
        for ci in range(3,14):
            sc(ws,r,ci,None,sz=10,bg=bg,ha="right",fmt=JPY)
        sc(ws,r,14,f"=IF(C{r}+D{r}+E{r}+F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r}+M{r}=0,\"\",C{r}+D{r}+E{r}+F{r}+G{r}+H{r}+I{r}+J{r}+K{r}+L{r}+M{r})",
              bold=True,sz=11,bg=bg,ha="right",fmt=JPY)
        sc(ws,r,15,f"=IF(N{r}=\"\",\"\",ROUND(-N{r}/11,0))",sz=10,bg=bg,ha="right",fmt=JPY)
        c16=ws.cell(row=r,column=16)
        c16.value=f"=IF(N{r}=\"\",\"\",N{r}+O{r})"; c16.number_format=JPY
        c16.font=fnt(bold=True,sz=11); c16.fill=fill(bg)
        c16.alignment=aln(h="right"); c16.border=bdr()

    # 色スケール（合計列）
    end_r = last + 12
    ws.conditional_formatting.add(f"N5:N{end_r}",
        ColorScaleRule(start_type="min",start_color="FFF9C4",
                       end_type="max",  end_color="1A5FC8"))

# ════════════════════════════════════════════════════════════════
#  SHEET B: 科目-案件マッピング
# ════════════════════════════════════════════════════════════════
# 配分設定: 科目→案件の%
# C001=東急, C002=都営, C003=一般
DEFAULT_MAPPING = {
    "612-1": (0,   0,  100),  # 運送本社 → 一般事業者
    "612-2": (0,  100,   0),  # 入札公共 → 都営地下鉄
    "613-1": (20,  10,  70),  # 車輌     → 各案件
    "613-2": (0,    0,   0),  # KC       → 別管理
    "614-1": (0,    0, 100),  # コンテナ → 一般
    "614-2": (0,    0, 100),  # ENDO     → 一般
    "614-3": (0,    0, 100),  # 現金     → 一般
    "615-1": (60,   0,  40),  # 板橋     → 東急60%・一般40% ★調整可
    "615-2": (0,    0, 100),  # 新横浜   → 一般
    "615-3": (0,    0, 100),  # 現金2    → 一般
    "615-4": (0,    0, 100),  # 有価物   → 一般
}

def build_mapping(ws):
    ws.title = "⑥科目-案件マッピング"
    ws.sheet_view.showGridLines = False

    title_row(ws, "科目-案件マッピング設定　※配分率(%)を調整することで③採算入力の売上が変わります", 9)

    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:I2")
    n = ws.cell(row=2,column=1,
        value="★ C001/C002/C003 の合計が100%になるよう配分を設定してください（KCのみ別管理で0%可）")
    n.font=fnt(sz=10,color="8B4500"); n.fill=fill(C_YELLOW)
    n.alignment=aln(h="left"); n.border=bdr()

    sec_hdr(ws,3,[("科目情報",1,3,C_NAVY),
                  ("案件別配分率（%）",4,6,"1A6B2E"),
                  ("確認",7,9,"1565A8")],h=22)

    hdrs=[("科目コード",13),("科目名",22),("性質",14),
          ("C001\n東急グループ%",14),("C002\n都営地下鉄%",14),("C003\n一般事業者%",14),
          ("合計%\n(100が正常)",13),("直近月売上\n(参照)",16),("直近月各案件\n貢献額(参照)",16)]
    col_hdrs(ws,4,[(h,w,C_BLUE) for h,w in hdrs],h=36)

    nature = {
        "612-1":"民間運送","612-2":"公共入札","613-1":"車輌","613-2":"KC売上",
        "614-1":"コンテナ","614-2":"ENDO","614-3":"現金","615-1":"板橋営業所",
        "615-2":"新横浜営業所","615-3":"現金2","615-4":"有価物",
    }

    # Get latest month data reference (will be linked to 売上原票)
    for r, (code, name) in enumerate(KAMOKU_CODES, 5):
        ws.row_dimensions[r].height = 24
        c001p, c002p, c003p = DEFAULT_MAPPING.get(code,(0,0,0))
        bg = C_LGRAY if r%2==0 else C_WHITE

        sc(ws,r,1,code, bold=True,sz=11,bg=bg,ha="center")
        sc(ws,r,2,name, sz=11,bg=bg)
        sc(ws,r,3,nature.get(code,""), sz=10,bg=bg,ha="center")
        # 入力セル（配分率）- 緑背景で入力可能を示す
        sc(ws,r,4,c001p,sz=12,bold=True,bg="E8F5E9",ha="center",fmt="0")
        sc(ws,r,5,c002p,sz=12,bold=True,bg="F3E5F5",ha="center",fmt="0")
        sc(ws,r,6,c003p,sz=12,bold=True,bg="E0F2F1",ha="center",fmt="0")
        # 合計確認
        c7=ws.cell(row=r,column=7)
        c7.value=f"=D{r}+E{r}+F{r}"
        c7.number_format="0"; c7.font=fnt(bold=True,sz=12)
        c7.fill=fill("C8F7D4"); c7.alignment=aln(h="center"); c7.border=bdr()
        # 条件付き書式（合計が100以外は赤）
        ws.conditional_formatting.add(f"G{r}:G{r}",
            FormulaRule(formula=[f"G{r}<>100"],fill=fill(C_RED),
                        font=Font(bold=True,color="8B0000")))

        # 直近月売上（VLOOKUP from 売上原票 - 最終行）
        col_index = 3 + CODES.index(code)  # 売上原票での列インデックス
        c8=ws.cell(row=r,column=8)
        c8.value=f'=IFERROR(VLOOKUP(MAX(⑤売上原票!$A$5:$A$100),⑤売上原票!$A:$P,{col_index},0),"-")'
        c8.number_format="#,##0"; c8.font=fnt(sz=11)
        c8.fill=fill(bg); c8.alignment=aln(h="right"); c8.border=bdr()

        c9=ws.cell(row=r,column=9)
        c9.value=f"=IF(H{r}=\"-\",\"-\",H{r}*(D{r}+E{r}+F{r})/100)"
        c9.number_format="#,##0"; c9.font=fnt(sz=11)
        c9.fill=fill(bg); c9.alignment=aln(h="right"); c9.border=bdr()

    # 案件別合計
    last = 4 + len(KAMOKU_CODES)
    tr = last + 1
    ws.row_dimensions[tr].height = 26
    sc(ws,tr,1,"合計",bold=True,sz=12,bg=C_NAVY,fc=C_WHITE,ha="center")
    ws.merge_cells(f"A{tr}:C{tr}")
    for col,l in [(4,"D"),(5,"E"),(6,"F"),(8,"H"),(9,"I")]:
        c=ws.cell(row=tr,column=col)
        c.value=f"=SUM({l}5:{l}{last})"; c.number_format=("#,##0" if col in (8,9) else "0")
        c.font=fnt(bold=True,sz=12,color=C_WHITE); c.fill=fill(C_NAVY)
        c.alignment=aln(h="right"); c.border=bdr()
    sc(ws,tr,7,"",bold=True,sz=12,bg=C_NAVY,fc=C_WHITE,ha="center")

    # 凡例
    nr = tr + 2
    ws.merge_cells(f"A{nr}:I{nr}")
    note = ws.cell(row=nr,column=1,
        value="【配分の考え方の例】　615-1(板橋)：東急グループの拠点は板橋営業所担当のため東急60%・一般40%と設定。実際の売上構成に合わせて調整してください。")
    note.font=fnt(sz=9,color=C_DGRAY,name="Yu Gothic UI"); note.fill=fill(C_LGRAY)
    note.alignment=aln(h="left"); note.border=bdr()
    ws.row_dimensions[nr].height = 18

# ════════════════════════════════════════════════════════════════
#  SHEET C: 月次採算入力（VLOOKUP連携版）
# ════════════════════════════════════════════════════════════════
def build_input_v3(ws, sales_data):
    ws.title = "③月次採算入力"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "F5"

    title_row(ws, "月次採算入力シート（案件別）　★黄緑セル=売上原票から自動参照", 22)

    # 説明行
    ws.row_dimensions[2].height = 22
    ws.merge_cells("A2:V2")
    n=ws.cell(row=2,column=1,
        value="【売上の自動連携】 E列〜H列の売上は ⑤売上原票 と ⑥科目-案件マッピング を参照して自動計算します。"
              " 集計月(D列)に YYYY-MM 形式で入力してください。")
    n.font=fnt(sz=10,color="1A5F00"); n.fill=fill("E8F5E9")
    n.alignment=aln(h="left"); n.border=bdr()

    sec_hdr(ws,3,[
        ("案件情報",1,4,C_NAVY),
        ("売上（⑤⑥から自動参照）",5,9,"155A00"),
        ("原　価（入力欄）",10,17,"7B3F00"),
        ("採　算",18,22,C_GREEN[:6] if len(C_GREEN)==6 else "1A6B2E"),
    ],h=22)
    sec_hdr(ws,4,[
        ("案件ID",1,1,C_NAVY),("案件名",2,2,C_NAVY),("カテゴリ",3,3,C_NAVY),("集計月\n(YYYY-MM)",4,4,C_NAVY),
        ("●運送\n(612-1)",5,5,"155A00"),("●入札\n(612-2)",6,6,"155A00"),
        ("●板橋\n(615-1)",7,7,"155A00"),("●その他",8,8,"155A00"),("売上合計",9,9,"155A00"),
        ("人件費",10,10,"7B3F00"),("燃料費",11,11,"7B3F00"),("車両維持費",12,12,"7B3F00"),
        ("中間処理費",13,13,"7B3F00"),("搬入運搬費",14,14,"7B3F00"),
        ("電子M代行費",15,15,"7B3F00"),("管理費",16,16,"7B3F00"),("原価合計",17,17,"7B3F00"),
        ("粗利益",18,18,"1A6B2E"),("粗利率",19,19,"1A6B2E"),
        ("採算判定",20,20,"1A6B2E"),("前月比\n売上",21,21,"1A6B2E"),("備考",22,22,"1A6B2E"),
    ],h=36)

    cw=[9,22,14,13,  14,14,14,12,16,  14,12,12,13,12,12,11,14,  14,10,11,12,16]
    for i,w in enumerate(cw,1):
        ws.column_dimensions[get_column_letter(i)].width=w

    JPY="#,##0"; PCT="0.0%"

    # 案件・月のサンプル
    cases=[
        ("C001","東急グループ案件","東急グループ"),
        ("C002","都営地下鉄案件",  "都営地下鉄"),
        ("C003","一般事業者案件",  "一般事業者"),
    ]
    # 直近の実績月
    avail_months = sorted(sales_data.keys())[-4:]  # 直近4ヶ月
    cost_samples={
        ("C001","2025-05"):(680000,185000,95000,620000,85000,45000,120000),
        ("C001","2025-06"):(690000,195000,98000,635000,90000,47000,125000),
        ("C001","2025-07"):(705000,198000,100000,648000,92000,48000,128000),
        ("C001","2025-08"):(698000,190000,97000,638000,88000,46000,122000),
        ("C002","2025-05"):(415000,120000,65000,372000,42000,28000,82000),
        ("C002","2025-06"):(428000,128000,68000,385000,44000,30000,88000),
        ("C002","2025-07"):(445000,135000,72000,405000,46000,32000,92000),
        ("C002","2025-08"):(438000,130000,70000,395000,44000,31000,90000),
        ("C003","2025-05"):(1340000,472000,215000,1155000,160000,82000,205000),
        ("C003","2025-06"):(1365000,502000,228000,1205000,172000,88000,218000),
        ("C003","2025-07"):(1410000,530000,238000,1258000,180000,92000,228000),
        ("C003","2025-08"):(1395000,515000,232000,1235000,175000,90000,222000),
    }

    r = 5
    first_row_per_case = {}
    for cid, cname, cat in cases:
        first_row_per_case[cid] = r
        for ym in avail_months:
            ws.row_dimensions[r].height = 22
            bg = CASE_BG.get(cid, C_LGRAY)
            fg_c = CASE_FG.get(cid, C_BLUE)

            sc(ws,r,1,cid,  bold=True,sz=11,bg=bg,ha="center")
            sc(ws,r,2,cname,sz=11,bg=bg)
            c3=ws.cell(row=r,column=3,value=cat)
            c3.font=fnt(bold=True,sz=10,color=C_WHITE); c3.fill=fill(fg_c)
            c3.alignment=aln(h="center"); c3.border=bdr()
            sc(ws,r,4,ym,sz=11,bg=bg,ha="center")

            # 売上列（VLOOKUP自動参照）- 案件別に科目配分を使う
            # E列: 612-1 × C001の配分率 → この案件への貢献分
            # 式: =IFERROR(VLOOKUP(D{r}, ⑤売上原票!$A:$P, 3, 0) * VLOOKUP("C001", ⑥科目マッピング!... , ...) / 100, 0)
            # 簡略化: VLOOKUP(年月, 原票, 列) × (マッピング配分/100)
            # 科目列: 612-1=col3, 612-2=col4, 615-1=col10 (in 売上原票)

            # 配分率の参照: ⑥シートの D列(C001%), E列(C002%), F列(C003%)
            # ⑥シートの行5=612-1, 行6=612-2, ..., 行12=615-1
            case_col = {"C001": "D", "C002": "E", "C003": "F"}[cid]

            # 売上列の VLOOKUP 数式
            uriage_col = {"612-1":3,"612-2":4,"615-1":10}
            _case_pct_col = {"C001": 4, "C002": 5, "C003": 6}[cid]

            # E: 612-1 の当案件分
            e = ws.cell(row=r,column=5)
            e.value = (f'=IFERROR(VLOOKUP(D{r},⑤売上原票!$A:$P,3,0)'
                       f'*VLOOKUP("612-1",⑥科目マッピング!$A:$F,{4 if cid=="C001" else 5 if cid=="C002" else 6},0)/100,0)')
            e.number_format=JPY; e.font=fnt(sz=10,bold=False)
            e.fill=fill("E8F5E9"); e.alignment=aln(h="right"); e.border=bdr()

            # F: 612-2 の当案件分
            f_ = ws.cell(row=r,column=6)
            f_.value = (f'=IFERROR(VLOOKUP(D{r},⑤売上原票!$A:$P,4,0)'
                        f'*VLOOKUP("612-2",⑥科目マッピング!$A:$F,{4 if cid=="C001" else 5 if cid=="C002" else 6},0)/100,0)')
            f_.number_format=JPY; f_.font=fnt(sz=10)
            f_.fill=fill("E8F5E9"); f_.alignment=aln(h="right"); f_.border=bdr()

            # G: 615-1 の当案件分
            g_ = ws.cell(row=r,column=7)
            g_.value = (f'=IFERROR(VLOOKUP(D{r},⑤売上原票!$A:$P,10,0)'
                        f'*VLOOKUP("615-1",⑥科目マッピング!$A:$F,{4 if cid=="C001" else 5 if cid=="C002" else 6},0)/100,0)')
            g_.number_format=JPY; g_.font=fnt(sz=10)
            g_.fill=fill("E8F5E9"); g_.alignment=aln(h="right"); g_.border=bdr()

            # H: その他科目合計 × 配分率 (613,614,615-2〜4)
            others_col_formula = "+".join(
                [f'IFERROR(VLOOKUP(D{r},⑤売上原票!$A:$P,{ci},0)*VLOOKUP("{code}",⑥科目マッピング!$A:$F,{4 if cid=="C001" else 5 if cid=="C002" else 6},0)/100,0)'
                 for ci,code in [(5,"613-1"),(6,"613-2"),(7,"614-1"),(8,"614-2"),(9,"614-3"),(11,"615-2"),(12,"615-3"),(13,"615-4")]]
            )
            h_ = ws.cell(row=r,column=8)
            h_.value = f"={others_col_formula}"
            h_.number_format=JPY; h_.font=fnt(sz=10)
            h_.fill=fill("E8F5E9"); h_.alignment=aln(h="right"); h_.border=bdr()

            # I: 売上合計
            i_ = ws.cell(row=r,column=9)
            i_.value = f"=E{r}+F{r}+G{r}+H{r}"
            i_.number_format=JPY; i_.font=fnt(bold=True,sz=12)
            i_.fill=fill("C8F7D4"); i_.alignment=aln(h="right"); i_.border=bdr()

            # 原価列（入力）
            costs = cost_samples.get((cid,ym),(0,0,0,0,0,0,0))
            for ci, v in enumerate(costs, 10):
                sc(ws,r,ci,v,sz=10,bg=C_YELLOW,ha="right",fmt=JPY)

            # Q: 原価合計
            q_ = ws.cell(row=r,column=17)
            q_.value = f"=J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r}"
            q_.number_format=JPY; q_.font=fnt(bold=True,sz=11)
            q_.fill=fill(C_YELLOW); q_.alignment=aln(h="right"); q_.border=bdr()

            # R: 粗利
            rr_ = ws.cell(row=r,column=18)
            rr_.value = f"=I{r}-Q{r}"
            rr_.number_format=JPY; rr_.font=fnt(bold=True,sz=12)
            rr_.fill=fill(C_GRN_LT); rr_.alignment=aln(h="right"); rr_.border=bdr()

            # S: 粗利率
            s_ = ws.cell(row=r,column=19)
            s_.value = f"=IF(I{r}<>0,R{r}/I{r},0)"
            s_.number_format=PCT; s_.font=fnt(bold=True,sz=11)
            s_.fill=fill(C_GRN_LT); s_.alignment=aln(h="center"); s_.border=bdr()

            # T: 採算判定
            t_ = ws.cell(row=r,column=20)
            t_.value = f'=IF(S{r}>=0.25,"◎ 優良",IF(S{r}>=0.15,"○ 良好",IF(S{r}>=0.05,"△ 要注意","× 赤字")))'
            t_.font=fnt(bold=True,sz=10); t_.fill=fill(C_GRN_LT)
            t_.alignment=aln(h="center"); t_.border=bdr()

            # U: 前月比 売上
            if r == first_row_per_case[cid]:
                sc(ws,r,21,"—",sz=10,bg=bg,ha="center")
            else:
                u_=ws.cell(row=r,column=21)
                u_.value=f"=IF(I{r-1}<>0,I{r}/I{r-1}-1,\"\")"
                u_.number_format="+0.0%;-0.0%;0.0%"; u_.font=fnt(sz=10)
                u_.fill=fill(bg); u_.alignment=aln(h="center"); u_.border=bdr()

            sc(ws,r,22,"",sz=10,bg=bg)
            r += 1

    # 追加入力行（20行）
    for _ in range(20):
        ws.row_dimensions[r].height = 22
        bg = C_LGRAY if r%2==0 else C_WHITE
        for c in range(1,23):
            sc(ws,r,c,None,sz=10,bg=bg,
               ha="center" if c in (1,3,4,20) else "right" if c>=5 else "left")
        # 売上 VLOOKUP（汎用・合計列のみ）
        for c,expr in [
            (9, f"=E{r}+F{r}+G{r}+H{r}"),
            (17,f"=IF(J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r}=0,\"\",J{r}+K{r}+L{r}+M{r}+N{r}+O{r}+P{r})"),
            (18,f'=IF(OR(I{r}="",I{r}=0),"",I{r}-Q{r})'),
            (19,f'=IF(OR(I{r}="",I{r}=0),"",R{r}/I{r})'),
            (20,f'=IF(S{r}="","",IF(S{r}>=0.25,"◎ 優良",IF(S{r}>=0.15,"○ 良好",IF(S{r}>=0.05,"△ 要注意","× 赤字"))))'),
        ]:
            cx=ws.cell(row=r,column=c); cx.value=expr
            cx.number_format=(PCT if c==19 else JPY if c in(9,17,18) else "@")
            cx.font=fnt(bold=c in(9,17,18,19),sz=10)
            cx.fill=fill("E8F5E9" if c in(5,6,7,8) else C_YELLOW if 10<=c<=16 else
                         C_GRN_LT if c in(17,18,19,20) else bg)
            cx.alignment=aln(h="center" if c==20 else "right")
            cx.border=bdr()
        r += 1

    # 条件付き書式
    end_r = r-1
    ws.conditional_formatting.add(f"T5:T{end_r}",
        FormulaRule(formula=[f'T5="◎ 優良"'], fill=fill("C8F7D4"),font=Font(bold=True,color="145A32")))
    ws.conditional_formatting.add(f"T5:T{end_r}",
        FormulaRule(formula=[f'T5="○ 良好"'], fill=fill(C_GRN_LT),font=Font(bold=True,color="1A6B2E")))
    ws.conditional_formatting.add(f"T5:T{end_r}",
        FormulaRule(formula=[f'T5="△ 要注意"'],fill=fill(C_YELLOW), font=Font(bold=True,color="856404")))
    ws.conditional_formatting.add(f"T5:T{end_r}",
        FormulaRule(formula=[f'T5="× 赤字"'], fill=fill(C_RED),    font=Font(bold=True,color="8B0000")))
    ws.conditional_formatting.add(f"S5:S{end_r}",
        ColorScaleRule(start_type="num",start_value=0, start_color="FF6B6B",
                       mid_type="num", mid_value=0.15,mid_color="FFD93D",
                       end_type="num", end_value=0.3, end_color="6BCB77"))

# ════════════════════════════════════════════════════════════════
#  SHEET D: 案件マスタ（簡略版）
# ════════════════════════════════════════════════════════════════
def build_anken(ws, locations):
    ws.title = "①案件マスタ"
    ws.sheet_view.showGridLines = False
    title_row(ws,"民間廃棄物収集受託案件　案件マスタ",11)

    cats = defaultdict(lambda: {"count":0,"kg":0,"drivers":set()})
    for loc in locations:
        cat = loc["cat"]
        cats[cat]["count"] += 1
        cats[cat]["kg"] += loc["kg_week"]
        for d in str(loc["driver"]).split("・"):
            if d.strip(): cats[cat]["drivers"].add(d.strip())

    sec_hdr(ws,2,[("案件情報",1,7,C_NAVY),("拠点規模",8,10,"1565A8"),
                  ("売上参照（自動）",11,13,"155A00")],h=22)
    hdrs=[("案件ID",9),("案件名",22),("カテゴリ",14),("担当プレフィックス",18),
          ("ドライバー数",14),("ステータス",12),("備考",14),
          ("拠点数",10),("週間回収量(kg)",14),("月間回収量(kg)",14),
          ("直近月売上\n(合計)",16),("直近月\n原価(入力)",14),("直近月\n粗利率",12)]
    col_hdrs(ws,3,[(h,w,C_BLUE) for h,w in hdrs],h=36)

    cases=[("C001","東急グループ案件","東急グループ","TK-"),
           ("C002","都営地下鉄案件",  "都営地下鉄","MT-"),
           ("C003","一般事業者案件",  "一般事業者","GN-")]

    for r,(cid,cname,cat,pfx) in enumerate(cases,4):
        ws.row_dimensions[r].height = 28
        d = cats[cat]
        bg = CASE_BG.get(cid,C_LGRAY)
        fg_c = CASE_FG.get(cid,C_BLUE)

        sc(ws,r,1,cid,  bold=True,sz=11,bg=bg,ha="center")
        sc(ws,r,2,cname,bold=True,sz=11,bg=bg)
        c3=ws.cell(row=r,column=3,value=cat)
        c3.font=fnt(bold=True,sz=11,color=C_WHITE); c3.fill=fill(fg_c)
        c3.alignment=aln(h="center"); c3.border=bdr()
        sc(ws,r,4,pfx,   sz=11,bg=bg,ha="center")
        sc(ws,r,5,len(d["drivers"]),sz=11,bg=bg,ha="center",fmt="#,##0")
        sc(ws,r,6,"稼働中",sz=11,bg=bg,ha="center")
        sc(ws,r,7,"",sz=10,bg=bg)
        sc(ws,r,8, d["count"],bold=True,sz=12,bg=bg,ha="center",fmt="#,##0")
        sc(ws,r,9, d["kg"],   bold=True,sz=12,bg=bg,ha="right", fmt="#,##0")
        c10=ws.cell(row=r,column=10)
        c10.value=f"=I{r}*4.33"; c10.number_format="#,##0"
        c10.font=fnt(bold=True,sz=12); c10.fill=fill(bg)
        c10.alignment=aln(h="right"); c10.border=bdr()

        # VLOOKUP: 売上原票の最新月を参照して案件別売上を表示
        case_col_idx = {"C001":4,"C002":5,"C003":6}[cid]
        c11=ws.cell(row=r,column=11)
        c11.value=(f'=IFERROR(SUMPRODUCT('
                   f'IFERROR(VLOOKUP(MAX(⑤売上原票!$A$5:$A$100),⑤売上原票!$A:$P,{{3,4,5,6,7,8,9,10,11,12,13}},0),0)'
                   f'*IFERROR(INDEX(⑥科目マッピング!$D$5:$F$15,,{case_col_idx-3}),0)/100)'
                   f',"算出中")')
        c11.number_format="#,##0"; c11.font=fnt(bold=True,sz=12)
        c11.fill=fill("E8F5E9"); c11.alignment=aln(h="right"); c11.border=bdr()
        sc(ws,r,12,0,sz=11,bg="E8F5E9",ha="right",fmt="#,##0")
        c13=ws.cell(row=r,column=13)
        c13.value=f'=IF(AND(K{r}<>"算出中",L{r}<>0),(K{r}-L{r})/K{r},"")'
        c13.number_format="0.0%"; c13.font=fnt(bold=True,sz=12)
        c13.fill=fill("E8F5E9"); c13.alignment=aln(h="center"); c13.border=bdr()

# ════════════════════════════════════════════════════════════════
#  BUILD
# ════════════════════════════════════════════════════════════════
def build():
    print("Loading data...")
    sales_data = load_all_sales()
    locations  = load_locations()
    drivers    = load_drivers()
    print(f"  売上データ: {len(sales_data)}ヶ月分")
    print(f"  拠点数: {len(locations)}")

    wb = Workbook()
    wb.remove(wb.active)

    ws_anken  = wb.create_sheet()
    ws_kyoten = wb.create_sheet()
    ws_input  = wb.create_sheet()
    ws_uriage = wb.create_sheet()
    ws_map    = wb.create_sheet()
    ws_dash   = wb.create_sheet()
    ws_driver = wb.create_sheet()
    ws_conv   = wb.create_sheet()

    # ① 案件マスタ
    build_anken(ws_anken, locations)

    # ② 拠点マスター（v2から流用・簡略版）
    ws_kyoten.title = "②拠点マスター"
    ws_kyoten.sheet_view.showGridLines = False
    title_row(ws_kyoten, f"拠点マスター（{len(locations)}拠点）", 11)
    sec_hdr(ws_kyoten,2,[("拠点基本情報",1,8,C_NAVY)],h=22)
    hdrs2=[("拠点コード",11),("拠点名",30),("カテゴリ",14),("住所",30),
           ("回収曜日",14),("担当ドライバー",20),("廃棄物種別",22),("週間回収量(kg)",14)]
    col_hdrs(ws_kyoten,3,[(h,w,C_BLUE) for h,w in hdrs2],h=22)
    cat_bg={"東急グループ":CASE_BG["C001"],"都営地下鉄":CASE_BG["C002"],"一般事業者":CASE_BG["C003"]}
    for ri,loc in enumerate(locations,4):
        ws_kyoten.row_dimensions[ri].height=18
        bg=cat_bg.get(loc["cat"],C_WHITE)
        sc(ws_kyoten,ri,1,loc["code"],  sz=10,bg=bg,ha="center")
        sc(ws_kyoten,ri,2,loc["name"],  sz=10,bg=bg,wrap=True)
        c3x=ws_kyoten.cell(row=ri,column=3,value=loc["cat"])
        c3x.font=fnt(sz=10,color=C_WHITE)
        c3x.fill=fill(CASE_FG.get({"東急グループ":"C001","都営地下鉄":"C002","一般事業者":"C003"}.get(loc["cat"],""),C_DGRAY))
        c3x.alignment=aln(h="center"); c3x.border=bdr()
        sc(ws_kyoten,ri,4,loc["addr"],  sz=9, bg=bg,wrap=True)
        sc(ws_kyoten,ri,5,loc["day"],   sz=9, bg=bg,ha="center")
        sc(ws_kyoten,ri,6,loc["driver"],sz=9, bg=bg,wrap=True)
        sc(ws_kyoten,ri,7,loc["wtype"], sz=9, bg=bg,wrap=True)
        sc(ws_kyoten,ri,8,loc["kg_week"],sz=10,bold=True,bg=bg,ha="right",fmt="#,##0")

    # ③ 月次採算入力
    build_input_v3(ws_input, sales_data)

    # ④ 案件別月次集計（簡略）
    ws_dash2 = wb.create_sheet()   # placeholder

    # ⑤ 売上原票
    build_uriagehyo(ws_uriage, sales_data)

    # ⑥ 科目-案件マッピング
    build_mapping(ws_map)

    # ⑦ ドライバー別（簡略）
    ws_driver.title = "⑦ドライバー別"
    ws_driver.sheet_view.showGridLines = False
    title_row(ws_driver,"ドライバー別回収量サマリー",7)
    dh=[("ドライバー名",14),("担当拠点数",13),("週間回収量(kg)",14),
        ("月間回収量(kg推算)",16),("1拠点あたり(kg)",14),("備考",16)]
    col_hdrs(ws_driver,2,[(h,w,C_BLUE) for h,w in dh],h=22)
    for ri,dr in enumerate(drivers,3):
        bg=C_LGRAY if ri%2==0 else C_WHITE
        ws_driver.row_dimensions[ri].height=22
        for ci,v in enumerate(dr[:5],1):
            try: fv=float(v) if v else 0
            except: fv=0
            sc(ws_driver,ri,ci,v,sz=11,bg=bg,ha="right" if ci>=2 else "left",
               fmt="#,##0" if ci>=2 else None)
        sc(ws_driver,ri,6,"",sz=10,bg=bg)

    # ⑧ ダッシュボード（簡略）
    ws_dash2.title = "⑧採算ダッシュボード"
    ws_dash2.sheet_view.showGridLines = False
    title_row(ws_dash2,"採算ダッシュボード（③月次採算入力 から自動集計）",8)
    note=ws_dash2.cell(row=3,column=1,
        value="③月次採算入力シートの入力値が蓄積されると、こちらに集計が表示されます。"
              " 現状は③シートの直近月データをご参照ください。")
    note.font=fnt(sz=13,color=C_NAVY); note.alignment=aln(h="left")
    ws_dash2.row_dimensions[3].height=30

    # ⑨ 換算テーブル
    wb_m=openpyxl.load_workbook(SRC_MASTER,data_only=True)
    ws_cv=wb_m["②換算テーブル"]
    ws_conv.title="⑨換算テーブル"
    ws_conv.sheet_view.showGridLines=False
    title_row(ws_conv,"品目×袋サイズ 重量換算・CO₂係数テーブル",6)
    ch=[("品目",16),("容器・サイズ",20),("1単位あたり重量(kg)",20),("備考",32),("CO₂排出係数",16)]
    col_hdrs(ws_conv,2,[(h,w,C_BLUE) for h,w in ch],h=28)
    for ri,row in enumerate(ws_cv.iter_rows(min_row=3,values_only=True),3):
        if row[0] and not str(row[0]).startswith("【"):
            bg=C_LGRAY if ri%2==0 else C_WHITE
            ws_conv.row_dimensions[ri].height=20
            for ci,v in enumerate(row[:5],1):
                sc(ws_conv,ri,ci,v,sz=10,bg=bg,
                   ha="right" if ci in(3,5) else "left",
                   fmt=("#,##0.0" if ci==3 else "0.00000" if ci==5 else None))

    # Tab colors
    for ws_,color in [(ws_anken,C_NAVY),(ws_kyoten,"1565A8"),(ws_input,"155A00"),
                      (ws_uriage,"E07B00"),(ws_map,"1A6B2E"),(ws_dash2,"FF6B00"),
                      (ws_driver,"5B2D8E"),(ws_conv,C_DGRAY)]:
        ws_.sheet_properties.tabColor = color

    # 削除：不要な placeholder
    wb.remove(ws_dash2)
    ws_dash2_new = wb.create_sheet("⑧採算ダッシュボード")
    ws_dash2_new.sheet_view.showGridLines = False
    title_row(ws_dash2_new,"採算ダッシュボード",8,bg=C_NAVY)
    n=ws_dash2_new.cell(row=3,column=1,
        value="③月次採算入力の実績が蓄積されると、ここに集計が表示されます。現状は③シートをご参照ください。")
    n.font=fnt(sz=13); n.alignment=aln(h="left")
    ws_dash2_new.sheet_properties.tabColor = "FF6B00"

    out="/home/user/sacure/sacure_採算管理_v3_VLOOKUP連携.xlsx"
    wb.save(out)
    print(f"Saved: {out}")
    print("シート構成:")
    for ws_ in wb.worksheets:
        print(f"  {ws_.title}")

if __name__=="__main__":
    build()
